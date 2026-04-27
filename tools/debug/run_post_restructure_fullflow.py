#!/usr/bin/env python
from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl
import serial

SCRIPT_DIR = Path(__file__).resolve().parent
AUDIO_DIR = SCRIPT_DIR.parent / "audio"
CASES_DIR = SCRIPT_DIR.parent / "cases"
if str(AUDIO_DIR) not in sys.path:
    sys.path.insert(0, str(AUDIO_DIR))

from fan_validation_helper import ensure_cached_tts, update_response_state
from listenai_play_repo import resolve_listenai_play

ROOT = Path(__file__).resolve().parents[2]
REQ_DIR = Path(os.environ.get("TRISOLARIS_REQ_DIR", ROOT / "项目需求" / "CSK5062小度风扇需求")).expanduser().resolve()
DELIVERABLE_ROOT = ROOT / "deliverables" / "csk5062_xiaodu_fan"
REPORT_ROOT = DELIVERABLE_ROOT / "reports"
PLAN_PATH = DELIVERABLE_ROOT / "plan" / "测试方案.md"
CASE_MD_PATH = DELIVERABLE_ROOT / "archive" / "测试用例-正式版.md"
CASE_XLSX_PATH = DELIVERABLE_ROOT / "cases" / "测试用例-正式版.xlsx"
FIRMWARE_OVERRIDE = os.environ.get("TRISOLARIS_FIRMWARE_BIN", "").strip()
FIRMWARE_PATH = Path(FIRMWARE_OVERRIDE).expanduser().resolve() if FIRMWARE_OVERRIDE else REQ_DIR / "fw-csk5062_xiaodu_fan-v1.0.0.bin"
GENERATE_ASSETS_SCRIPT = CASES_DIR / "generate_formal_assets.py"
BUNDLE_TAG = os.environ.get("TRISOLARIS_BUNDLE_TAG", "").strip()

DEFAULT_DEVICE_KEY = "VID_8765&PID_5678:8_804B35B_1_0000"
DEFAULT_LOG_PORT = "COM38"
DEFAULT_LOG_BAUD = 115200
DEFAULT_PROTO_PORT = "COM36"
DEFAULT_PROTO_BAUD = 9600
DEFAULT_CTRL_PORT = "COM39"
DEFAULT_CTRL_BAUD = 115200
DEFAULT_BURN_BAUD = 1500000


def env_text(name: str, default: str = "") -> str:
    value = os.environ.get(name, "").strip()
    return value or default


def env_int(name: str, default: int) -> int:
    value = os.environ.get(name, "").strip()
    if not value:
        return default
    try:
        return int(value)
    except ValueError as exc:
        raise RuntimeError(f"Environment variable {name} must be an integer, got: {value}") from exc


LOG_PORT = env_text("TRISOLARIS_LOG_PORT", DEFAULT_LOG_PORT)
LOG_BAUD = env_int("TRISOLARIS_LOG_BAUD", DEFAULT_LOG_BAUD)
PROTO_PORT = env_text("TRISOLARIS_PROTO_PORT", DEFAULT_PROTO_PORT)
PROTO_BAUD = env_int("TRISOLARIS_PROTO_BAUD", DEFAULT_PROTO_BAUD)
CTRL_PORT = env_text("TRISOLARIS_CTRL_PORT", DEFAULT_CTRL_PORT)
CTRL_BAUD = env_int("TRISOLARIS_CTRL_BAUD", DEFAULT_CTRL_BAUD)
BURN_PORT = env_text("TRISOLARIS_BURN_PORT", LOG_PORT)
BURN_BAUD = env_int("TRISOLARIS_BURN_BAUD", DEFAULT_BURN_BAUD)
DEVICE_KEY_OVERRIDE = os.environ.get("TRISOLARIS_DEVICE_KEY", "").strip()

DEFAULT_BETWEEN_MAX_WAIT_S = 4.5
DEFAULT_BETWEEN_MIN_WAIT_S = 0.8
DEFAULT_QUIET_WINDOW_S = 0.6
DEFAULT_POST_WAIT_S = 4.0


def iso_now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def sanitize_name(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("_")
    return cleaned or "step"


def decode_text(data: bytes) -> str:
    return data.decode("utf-8", errors="replace")


def text_has_any(text: str, markers: Iterable[str]) -> bool:
    return any(marker in text for marker in markers)


def format_proto_log(data: bytes) -> str:
    if not data:
        return ""
    lines = []
    for index in range(0, len(data), 4):
        lines.append(data[index : index + 4].hex(" ").upper())
    return "\n".join(lines) + "\n"


def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def scan_render_device_keys(listenai_play: Path) -> list[str]:
    try:
        completed = subprocess.run(
            [sys.executable, str(listenai_play), "scan"],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            check=False,
        )
    except OSError:
        return []
    if completed.returncode != 0:
        return []
    keys: list[str] = []
    for raw_line in completed.stdout.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("Direction "):
            continue
        parts = line.split()
        if len(parts) >= 2 and parts[0] == "Render":
            keys.append(parts[1])
    return keys


def resolve_playback_device_key(listenai_play: Path) -> str:
    if DEVICE_KEY_OVERRIDE:
        return DEVICE_KEY_OVERRIDE
    keys = scan_render_device_keys(listenai_play)
    if DEFAULT_DEVICE_KEY in keys:
        return DEFAULT_DEVICE_KEY
    if len(keys) == 1:
        return keys[0]
    if len(keys) > 1:
        raise RuntimeError(
            "Multiple render device keys detected; set TRISOLARIS_DEVICE_KEY explicitly: "
            + ", ".join(keys)
        )
    return DEFAULT_DEVICE_KEY


def is_yes(value: str) -> bool:
    return value.strip() in {"是", "支持", "保存", "需要", "true", "True", "YES", "Yes", "yes", "1"}


def parse_requirement_markdown(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8")

    def expect_int(pattern: str) -> int:
        match = re.search(pattern, text)
        if not match:
            raise RuntimeError(f"Unable to parse requirement field with pattern: {pattern}")
        return int(match.group(1))

    def expect_text(pattern: str) -> str:
        match = re.search(pattern, text)
        if not match:
            raise RuntimeError(f"Unable to parse requirement field with pattern: {pattern}")
        return match.group(1).strip()

    return {
        "wake_timeout_s": expect_int(r"唤醒时长:\s*(\d+)s"),
        "volume_steps": expect_int(r"音量档位:\s*(\d+)"),
        "default_volume": expect_int(r"初始化默认音量:\s*(\d+)"),
        "mic_analog_gain_db": expect_int(r"mic模拟增益:\s*(\d+)"),
        "mic_digital_gain_db": expect_int(r"mic数字增益:\s*(\d+)"),
        "proto_baud": expect_int(r"协议串口:\s*UART1、波特率(\d+)"),
        "log_baud": expect_int(r"日志串口:\s*UART0、波特率(\d+)"),
        "wake_power_save_raw": expect_text(r"唤醒词掉电保存:\s*([^\n]+)"),
        "volume_power_save_raw": expect_text(r"音量掉电保存:\s*([^\n]+)"),
    }


def load_word_table(path: Path) -> dict[str, dict[str, str]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[1]]
    items: dict[str, dict[str, str]] = {}
    headers = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        record = {str(headers[idx]): ("" if row[idx] is None else str(row[idx])) for idx in range(len(headers))}
        key = record.get("语义(最小功能词)") or record.get("功能类型")
        if key:
            items[key] = record
    return items


def load_voice_reg_config(path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows: list[tuple[str, str, Any]] = []
    current_category = ""
    for row in sheet.iter_rows(values_only=True):
        if not row or row[2] is None:
            continue
        if row[1]:
            current_category = str(row[1]).strip()
        key = str(row[2]).strip()
        value = row[3]
        rows.append((current_category, key, value))

    def find_value(category: str, key: str) -> Any:
        for item_category, item_key, item_value in rows:
            if item_category == category and item_key == key:
                return item_value
        raise RuntimeError(f"Unable to find voice registration config: {category} / {key}")

    return {
        "command_mode": str(find_value("自学习种类及模式", "命令词学习模式")),
        "wakeup_repeat_count": int(find_value("自学习唤醒词参数", "自学习时每个词需说几遍")),
        "wakeup_word_max": int(find_value("自学习唤醒词参数", "自学习唤醒词字数上限")),
        "wakeup_word_min": int(find_value("自学习唤醒词参数", "自学习唤醒词字数下限")),
        "wakeup_template_count": int(find_value("自学习唤醒词参数", "自学习唤醒词模板数")),
        "wakeup_retry_count": int(find_value("自学习唤醒词参数", "唤醒词学习失败重试次数")),
        "command_repeat_count": int(find_value("自学习命令词参数", "自学习时每个词需说几遍")),
        "command_word_max": int(find_value("自学习命令词参数", "自学习命令词字数上限")),
        "command_word_min": int(find_value("自学习命令词参数", "自学习命令词字数下限")),
        "command_template_count": int(find_value("自学习命令词参数", "自学习命令词模板数")),
        "command_retry_count": int(find_value("自学习命令词参数", "命令词学习失败重试次数")),
    }


def load_project_spec() -> dict[str, Any]:
    requirements = parse_requirement_markdown(REQ_DIR / "需求文档.md")
    words = load_word_table(REQ_DIR / "词条处理.xlsx")
    voice_reg = load_voice_reg_config(REQ_DIR / "语音注册功能.xlsx")
    return {
        "requirements": requirements,
        "words": words,
        "voice_reg": voice_reg,
    }


def parse_boot_config(log_text: str) -> dict[str, int | str]:
    config: dict[str, int | str] = {}
    in_block = False
    for raw_line in log_text.splitlines():
        line = raw_line.replace("\x1b", "")
        if "Running Config" in line:
            in_block = True
            continue
        if not in_block:
            continue
        if "==========================" in line:
            if config:
                break
            continue
        match = re.match(r"\s*([A-Za-z][A-Za-z0-9]+)\s*:\s*([^\s]+)", line)
        if not match:
            continue
        key, value = match.groups()
        if value.isdigit():
            config[key] = int(value)
        else:
            config[key] = value
    return config


def parse_mic_gain(log_text: str) -> dict[str, int]:
    match = re.search(r"AGAIN=(\d+)dB.*?DGAIN=\s*(\d+)dB", log_text, re.S)
    if not match:
        return {}
    return {
        "analog_gain_db": int(match.group(1)),
        "digital_gain_db": int(match.group(2)),
    }


def proto_frames_from_hex(proto_hex: str) -> list[str]:
    compact = proto_hex.replace("\n", " ").strip()
    if not compact:
        return []
    tokens = compact.split()
    frames = []
    for index in range(0, len(tokens), 4):
        frame = tokens[index : index + 4]
        if len(frame) == 4:
            frames.append(" ".join(frame))
    return frames


def evidence_has_frames(evidence: StepEvidence, expected_frames: list[str]) -> bool:
    frames = proto_frames_from_hex(evidence.proto_hex)
    cursor = 0
    for frame in expected_frames:
        try:
            cursor = frames.index(frame, cursor) + 1
        except ValueError:
            return False
    return True


def evidence_has_frame(evidence: StepEvidence, expected_frame: str) -> bool:
    return expected_frame in proto_frames_from_hex(evidence.proto_hex)


def count_occurrences(text: str, marker: str) -> int:
    return text.count(marker)


def extract_volume_values(log_text: str) -> list[int]:
    return [int(item) for item in re.findall(r"refresh config volume=(\d+)", log_text)]


def last_volume_value(log_text: str) -> int | None:
    values = extract_volume_values(log_text)
    return values[-1] if values else None


def extract_runtime_volume_levels(log_text: str) -> list[int]:
    return [int(item) for item in re.findall(r"mini player set vol\s*:\s*(\d+)", log_text)]


def last_runtime_volume_level(log_text: str) -> int | None:
    values = extract_runtime_volume_levels(log_text)
    return values[-1] if values else None


def ordered_unique(values: list[int]) -> list[int]:
    seen: set[int] = set()
    result: list[int] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


@dataclass
class StepEvidence:
    name: str
    step_dir: Path
    log_bytes: int
    proto_bytes: int
    log_text: str
    proto_hex: str
    detail: dict[str, Any]


class UntestableFirmware(RuntimeError):
    pass


class FullflowRunner:
    def __init__(self) -> None:
        self.spec = load_project_spec()
        self.stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = f"_{sanitize_name(BUNDLE_TAG)}" if BUNDLE_TAG else "_post_restructure_fullflow"
        self.bundle_dir = REPORT_ROOT / f"{self.stamp}{suffix}"
        self.static_dir = ensure_dir(self.bundle_dir / "01_static")
        self.burn_dir = ensure_dir(self.bundle_dir / "02_burn")
        self.exec_dir = ensure_dir(self.bundle_dir / "03_execution")
        self.steps_dir = ensure_dir(self.exec_dir / "steps")
        self.stream_dir = ensure_dir(self.exec_dir / "streams")
        self.case_results_path = self.exec_dir / "case_results.json"
        self.summary_md_path = self.exec_dir / "execution_summary.md"
        self.failure_analysis_path = self.exec_dir / "failure_analysis.md"
        self.testability_gate_path = self.exec_dir / "testability_gate.json"
        self.events_path = self.exec_dir / "events.jsonl"
        self.log_port: serial.Serial | None = None
        self.proto_port: serial.Serial | None = None
        self.listenai_play = resolve_listenai_play(update=False)
        self.device_key = resolve_playback_device_key(self.listenai_play)
        self.log_chunks = bytearray()
        self.proto_chunks = bytearray()
        self.case_results: list[dict[str, Any]] = []
        self.events: list[dict[str, Any]] = []
        self.testability_gate: dict[str, Any] = {}
        self.step_counter = 0

    def log_event(self, name: str, detail: dict[str, Any] | None = None) -> None:
        payload = {"at": iso_now(), "name": name, "detail": detail or {}}
        self.events.append(payload)
        with self.events_path.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(payload, ensure_ascii=False) + "\n")

    def add_case_result(
        self,
        case_id: str,
        module: str,
        status: str,
        summary: str,
        evidence: list[Path],
        detail: dict[str, Any] | None = None,
    ) -> None:
        item = {
            "case_id": case_id,
            "module": module,
            "status": status,
            "summary": summary,
            "evidence": [str(path.relative_to(ROOT)) for path in evidence],
            "detail": detail or {},
            "at": iso_now(),
        }
        self.case_results.append(item)
        self.log_event("case_result", item)
        print(f"[{status}] {case_id}: {summary}", flush=True)

    def prepare_static_assets(self) -> None:
        ensure_dir(self.static_dir / "requirement")
        ensure_dir(self.static_dir / "plan")
        ensure_dir(self.static_dir / "cases")
        shutil.copy2(PLAN_PATH, self.static_dir / "plan" / PLAN_PATH.name)
        shutil.copy2(CASE_MD_PATH, self.static_dir / "cases" / CASE_MD_PATH.name)
        if CASE_XLSX_PATH.exists():
            shutil.copy2(CASE_XLSX_PATH, self.static_dir / "cases" / CASE_XLSX_PATH.name)
        for item in [
            "需求文档.md",
            "词条处理.xlsx",
            "语音注册功能.xlsx",
            "tone.h",
        ]:
            src = REQ_DIR / item
            if src.exists():
                shutil.copy2(src, self.static_dir / "requirement" / src.name)
        if FIRMWARE_PATH.exists():
            shutil.copy2(FIRMWARE_PATH, self.static_dir / "requirement" / FIRMWARE_PATH.name)
        summary = {
            "firmware": str(FIRMWARE_PATH.relative_to(ROOT)),
            "device_key": self.device_key,
            "platform": "windows" if os.name == "nt" else "linux",
            "ports": {
                "proto": f"{PROTO_PORT}@{PROTO_BAUD}",
                "log": f"{LOG_PORT}@{LOG_BAUD}",
                "ctrl": f"{CTRL_PORT}@{CTRL_BAUD}",
                "burn": f"{BURN_PORT}@{BURN_BAUD}",
            },
            "generated_at": iso_now(),
        }
        (self.static_dir / "bundle_meta.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    def run_preburn_shell_command(self, name: str, command: str, *, capture_s: float, ready_wait_s: float) -> Path:
        step_dir = ensure_dir(self.burn_dir / name)
        raw_path = step_dir / "com38_raw.bin"
        text_path = step_dir / "com38_utf8.txt"
        meta_path = step_dir / "meta.json"
        chunks = bytearray()
        port = serial.Serial(LOG_PORT, baudrate=LOG_BAUD, timeout=0.05, write_timeout=0.5)
        try:
            port.reset_input_buffer()
            port.write((command + "\r\n").encode("ascii"))
            port.flush()
            deadline = time.time() + capture_s
            while time.time() < deadline:
                data = port.read(4096)
                if data:
                    chunks.extend(data)
                time.sleep(0.02)
        finally:
            port.close()
        raw_path.write_bytes(chunks)
        text = decode_text(bytes(chunks))
        text_path.write_text(text, encoding="utf-8")
        meta = {
            "command": command,
            "log_port": f"{LOG_PORT}@{LOG_BAUD}",
            "capture_s": capture_s,
            "ready_wait_s": ready_wait_s,
            "captured_bytes": len(chunks),
            "saved_at": iso_now(),
        }
        meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
        if ready_wait_s > 0:
            time.sleep(ready_wait_s)
        self.log_event("preburn_shell_command", {"name": name, "command": command, "step_dir": str(step_dir.relative_to(ROOT))})
        return step_dir

    def burn_firmware(self) -> None:
        self.log_event("burn_start", {"firmware": str(FIRMWARE_PATH)})
        preclear_dir = self.run_preburn_shell_command("00_preburn_config_clear", "config.clear", capture_s=3.0, ready_wait_s=1.0)
        prereboot_dir = self.run_preburn_shell_command("01_preburn_reboot", "reboot", capture_s=10.0, ready_wait_s=8.0)
        if os.name == "nt":
            bundle_dir = ROOT / "tools" / "burn_bundle" / "windows"
            cmd = [
                "powershell",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(ROOT / "tools" / "burn_bundle" / "run_fan_burn.ps1"),
                "-FirmwareBin",
                str(FIRMWARE_PATH),
                "-CtrlPort",
                CTRL_PORT,
                "-BurnPort",
                BURN_PORT,
                "-CtrlBaud",
                str(CTRL_BAUD),
                "-LogBaud",
                str(LOG_BAUD),
                "-BurnBaud",
                str(BURN_BAUD),
                "-PostPowerOnReadSeconds",
                "20",
            ]
        else:
            bundle_dir = ROOT / "tools" / "burn_bundle" / "linux"
            cmd = [
                "bash",
                str(ROOT / "tools" / "burn_bundle" / "run_fan_burn.sh"),
                "-FirmwareBin",
                str(FIRMWARE_PATH),
                "-CtrlPort",
                CTRL_PORT,
                "-BurnPort",
                BURN_PORT,
                "-CtrlBaud",
                str(CTRL_BAUD),
                "-LogBaud",
                str(LOG_BAUD),
                "-BurnBaud",
                str(BURN_BAUD),
                "-PostPowerOnReadSeconds",
                "20",
            ]
        completed = subprocess.run(cmd, cwd=ROOT, capture_output=True, text=True, encoding="utf-8", errors="replace", check=False)
        stdout_path = self.burn_dir / "run_output.txt"
        stderr_path = self.burn_dir / "run_error.txt"
        stdout_path.write_text(completed.stdout, encoding="utf-8")
        stderr_path.write_text(completed.stderr, encoding="utf-8")
        for name in ["burn.log", "burn_tool.log"]:
            src = bundle_dir / name
            if src.exists():
                shutil.copy2(src, self.burn_dir / name)
        burn_log_text = ""
        burn_log = self.burn_dir / "burn.log"
        if burn_log.exists():
            burn_log_text = burn_log.read_text(encoding="utf-8", errors="replace")
        burn_ok = completed.returncode == 0 and "Burn flow completed" in burn_log_text
        info = {
            "command": cmd,
            "returncode": completed.returncode,
            "burn_ok": burn_ok,
            "preburn_config_clear_dir": str(preclear_dir.relative_to(ROOT)),
            "preburn_reboot_dir": str(prereboot_dir.relative_to(ROOT)),
            "stdout_file": str(stdout_path.relative_to(ROOT)),
            "stderr_file": str(stderr_path.relative_to(ROOT)),
        }
        (self.burn_dir / "burn_meta.json").write_text(json.dumps(info, ensure_ascii=False, indent=2), encoding="utf-8")
        self.log_event("burn_finish", info)
        if not burn_ok:
            raise RuntimeError(f"Burn failed, see {self.burn_dir}")

    def open_ports(self) -> None:
        self.log_port = serial.Serial(LOG_PORT, baudrate=LOG_BAUD, timeout=0.05, write_timeout=0.5)
        self.proto_port = serial.Serial(PROTO_PORT, baudrate=PROTO_BAUD, timeout=0.05, write_timeout=0.5)
        self.log_port.reset_input_buffer()
        self.proto_port.reset_input_buffer()
        self.ensure_runtime_loglevel(capture_s=1.0, clear_buffers=True)
        self.log_event("ports_opened", {"log_port": LOG_PORT, "proto_port": PROTO_PORT})

    def close_ports(self) -> None:
        if self.proto_port:
            self.proto_port.close()
            self.proto_port = None
        if self.log_port:
            self.log_port.close()
            self.log_port = None
        self.log_event("ports_closed")

    def slice_offsets(self) -> tuple[int, int]:
        return (len(self.log_chunks), len(self.proto_chunks))

    def pump(self, duration_s: float, state: dict[str, Any] | None = None) -> None:
        assert self.log_port and self.proto_port
        deadline = time.time() + duration_s
        local_log = bytearray()
        while time.time() < deadline:
            log_data = self.log_port.read(4096)
            if log_data:
                self.log_chunks.extend(log_data)
                local_log.extend(log_data)
                if state is not None:
                    update_response_state(state, decode_text(local_log), time.time())
            proto_data = self.proto_port.read(4096)
            if proto_data:
                self.proto_chunks.extend(proto_data)
            time.sleep(0.02)

    def ensure_runtime_loglevel(self, capture_s: float = 1.0, clear_buffers: bool = False) -> None:
        assert self.log_port and self.proto_port
        if clear_buffers:
            self.log_port.reset_input_buffer()
            self.proto_port.reset_input_buffer()
        self.log_port.write(b"loglevel 4\r\n")
        self.log_port.flush()
        self.pump(capture_s)

    def wait_for_completion(self, initial_state: dict[str, Any]) -> dict[str, Any]:
        assert self.log_port and self.proto_port
        start = time.time()
        local_log = bytearray()
        state = {
            "saw_response": bool(initial_state.get("saw_response")),
            "saw_play_start": bool(initial_state.get("saw_play_start")),
            "saw_play_stop": bool(initial_state.get("saw_play_stop")),
            "last_data_at": initial_state.get("last_data_at"),
        }
        completion_reason: str | None = None
        while True:
            now = time.time()
            log_data = self.log_port.read(4096)
            if log_data:
                self.log_chunks.extend(log_data)
                local_log.extend(log_data)
                update_response_state(state, decode_text(local_log), now)
                if completion_reason == "quiet_after_response":
                    completion_reason = None
            proto_data = self.proto_port.read(4096)
            if proto_data:
                self.proto_chunks.extend(proto_data)
            elapsed = now - start
            last_data_at = state["last_data_at"]
            if state["saw_play_start"] and state["saw_play_stop"]:
                completion_reason = "play_stop"
            if completion_reason == "play_stop" and elapsed >= DEFAULT_BETWEEN_MIN_WAIT_S:
                return {"reason": completion_reason, "elapsed_s": round(elapsed, 3)}
            if state["saw_response"] and not state["saw_play_start"] and last_data_at is not None and (now - last_data_at) >= DEFAULT_QUIET_WINDOW_S:
                completion_reason = "quiet_after_response"
            if completion_reason == "quiet_after_response" and elapsed >= DEFAULT_BETWEEN_MIN_WAIT_S:
                return {"reason": completion_reason, "elapsed_s": round(elapsed, 3)}
            if elapsed >= DEFAULT_BETWEEN_MAX_WAIT_S:
                return {"reason": "max_wait", "elapsed_s": round(elapsed, 3)}
            time.sleep(0.02)

    def write_step_artifacts(self, name: str, start_offsets: tuple[int, int], extra_meta: dict[str, Any]) -> StepEvidence:
        self.step_counter += 1
        step_dir = ensure_dir(self.steps_dir / f"{self.step_counter:02d}_{sanitize_name(name)}")
        log_slice = bytes(self.log_chunks[start_offsets[0] :])
        proto_slice = bytes(self.proto_chunks[start_offsets[1] :])
        log_text = decode_text(log_slice)
        proto_hex = proto_slice.hex(" ").upper()
        (step_dir / "com38_raw.bin").write_bytes(log_slice)
        (step_dir / "com38_utf8.txt").write_text(log_text, encoding="utf-8")
        (step_dir / "com36_raw.bin").write_bytes(proto_slice)
        (step_dir / "com36_hex.txt").write_text(proto_hex, encoding="utf-8")
        detail = {"name": name, "log_bytes": len(log_slice), "proto_bytes": len(proto_slice), **extra_meta}
        (step_dir / "meta.json").write_text(json.dumps(detail, ensure_ascii=False, indent=2), encoding="utf-8")
        return StepEvidence(name=name, step_dir=step_dir, log_bytes=len(log_slice), proto_bytes=len(proto_slice), log_text=log_text, proto_hex=proto_hex, detail=detail)
    def run_voice_sequence(self, name: str, texts: list[str], post_wait_s: float = DEFAULT_POST_WAIT_S) -> StepEvidence:
        assert self.log_port and self.proto_port
        self.log_event("voice_sequence_start", {"name": name, "texts": texts})
        start_offsets = self.slice_offsets()
        audio_items = []
        playback_output = []
        between = []
        for index, text in enumerate(texts, start=1):
            audio_path, cached = ensure_cached_tts(text=text, voice="Microsoft Huihui Desktop", rate=0, label=f"{sanitize_name(name)}_{index}")
            audio_items.append({"text": text, "audio_file": str(audio_path), "cached": cached})
            cmd = [sys.executable, str(self.listenai_play), "play", "--audio-file", str(audio_path), "--device-key", self.device_key]
            proc = subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            response_state = {"saw_response": False, "saw_play_start": False, "saw_play_stop": False, "last_data_at": None}
            local_log = bytearray()
            while True:
                log_data = self.log_port.read(4096)
                if log_data:
                    self.log_chunks.extend(log_data)
                    local_log.extend(log_data)
                    update_response_state(response_state, decode_text(local_log), time.time())
                proto_data = self.proto_port.read(4096)
                if proto_data:
                    self.proto_chunks.extend(proto_data)
                if proc.poll() is not None:
                    break
                time.sleep(0.02)
            response_state["serial_bytes_during_playback"] = len(local_log)
            playback_output.append({"text": text, "audio_file": str(audio_path), "cached": cached, "output": [], "response_state": response_state})
            if index < len(texts):
                between.append(self.wait_for_completion(response_state))
        self.pump(post_wait_s)
        evidence = self.write_step_artifacts(name, start_offsets, {"mode": "voice_sequence", "texts": texts, "audio_items": audio_items, "playback_output": playback_output, "between_wait_result": between, "post_wait_s": post_wait_s})
        self.log_event("voice_sequence_finish", {"name": name, "log_bytes": evidence.log_bytes, "proto_bytes": evidence.proto_bytes})
        return evidence

    def run_protocol_step(self, name: str, payload_hex: str, post_wait_s: float = 4.0, pre_wait_s: float = 0.5) -> StepEvidence:
        assert self.proto_port
        self.log_event("protocol_start", {"name": name, "payload_hex": payload_hex})
        start_offsets = self.slice_offsets()
        self.pump(pre_wait_s)
        payload = bytes.fromhex(payload_hex.replace(" ", ""))
        self.proto_port.write(payload)
        self.proto_port.flush()
        self.pump(post_wait_s)
        evidence = self.write_step_artifacts(name, start_offsets, {"mode": "protocol", "payload_hex": payload_hex, "pre_wait_s": pre_wait_s, "post_wait_s": post_wait_s})
        self.log_event("protocol_finish", {"name": name, "log_bytes": evidence.log_bytes, "proto_bytes": evidence.proto_bytes})
        return evidence

    def run_shell_step(self, name: str, command: str, capture_s: float, ready_wait_s: float = 0.0) -> StepEvidence:
        assert self.log_port and self.proto_port
        self.log_event("shell_start", {"name": name, "command": command})
        start_offsets = self.slice_offsets()
        self.log_port.write((command + "\r\n").encode("ascii"))
        self.log_port.flush()
        self.pump(capture_s)
        if ready_wait_s > 0:
            self.pump(ready_wait_s)
        if command.strip().lower() == "reboot":
            self.ensure_runtime_loglevel(capture_s=1.0, clear_buffers=False)
        evidence = self.write_step_artifacts(name, start_offsets, {"mode": "shell", "command": command, "capture_s": capture_s, "ready_wait_s": ready_wait_s})
        self.log_event("shell_finish", {"name": name, "log_bytes": evidence.log_bytes, "proto_bytes": evidence.proto_bytes})
        return evidence

    def run_idle_wait_step(self, name: str, duration_s: float) -> StepEvidence:
        self.log_event("idle_wait_start", {"name": name, "duration_s": duration_s})
        start_offsets = self.slice_offsets()
        self.pump(duration_s)
        evidence = self.write_step_artifacts(name, start_offsets, {"mode": "idle_wait", "duration_s": duration_s})
        self.log_event("idle_wait_finish", {"name": name, "log_bytes": evidence.log_bytes, "proto_bytes": evidence.proto_bytes})
        return evidence

    def run_wake_timeout_probe(self, name: str, wake_text: str, wait_s: float = 24.0) -> StepEvidence:
        assert self.log_port and self.proto_port
        self.log_event("wake_timeout_probe_start", {"name": name, "wake_text": wake_text, "wait_s": wait_s})
        start_offsets = self.slice_offsets()
        audio_path, cached = ensure_cached_tts(text=wake_text, voice="Microsoft Huihui Desktop", rate=0, label=f"{sanitize_name(name)}_wake")
        cmd = [sys.executable, str(self.listenai_play), "play", "--audio-file", str(audio_path), "--device-key", self.device_key]
        probe_started_at = time.time()
        proc = subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        playback_output: list[str] = []
        local_log = bytearray()
        local_proto = bytearray()
        markers = {
            "wake_frame_s": None,
            "wake_keyword_s": None,
            "wake_play_start_s": None,
            "wake_play_stop_s": None,
            "timeout_marker_s": None,
            "mode_zero_s": None,
        }

        def update_markers(now: float) -> None:
            text = decode_text(local_log)
            elapsed = round(now - probe_started_at, 3)
            frames = proto_frames_from_hex(local_proto.hex(" ").upper())
            if markers["wake_frame_s"] is None and "A5 FA 01 BB" in frames:
                markers["wake_frame_s"] = elapsed
            if markers["wake_keyword_s"] is None and "keyword:xiao du xiao du" in text:
                markers["wake_keyword_s"] = elapsed
            if markers["wake_play_start_s"] is None and (markers["wake_frame_s"] is not None or markers["wake_keyword_s"] is not None) and ("play start" in text or "play id :" in text):
                markers["wake_play_start_s"] = elapsed
            if markers["wake_play_stop_s"] is None and (markers["wake_play_start_s"] is not None or markers["wake_frame_s"] is not None or markers["wake_keyword_s"] is not None) and "play stop" in text:
                markers["wake_play_stop_s"] = elapsed
            if markers["timeout_marker_s"] is None and "TIME_OUT" in text:
                markers["timeout_marker_s"] = elapsed
            if markers["mode_zero_s"] is None and "MODE=0" in text:
                markers["mode_zero_s"] = elapsed

        while True:
            now = time.time()
            log_data = self.log_port.read(4096)
            if log_data:
                self.log_chunks.extend(log_data)
                local_log.extend(log_data)
                update_markers(now)
            proto_data = self.proto_port.read(4096)
            if proto_data:
                self.proto_chunks.extend(proto_data)
                local_proto.extend(proto_data)
                update_markers(now)
            if proc.poll() is not None:
                break
            time.sleep(0.02)

        while time.time() - probe_started_at < wait_s:
            now = time.time()
            log_data = self.log_port.read(4096)
            if log_data:
                self.log_chunks.extend(log_data)
                local_log.extend(log_data)
                update_markers(now)
            proto_data = self.proto_port.read(4096)
            if proto_data:
                self.proto_chunks.extend(proto_data)
                local_proto.extend(proto_data)
                update_markers(now)
            if markers["mode_zero_s"] is not None:
                break
            time.sleep(0.02)

        response_end_s = markers["wake_play_stop_s"] or markers["wake_frame_s"] or markers["wake_keyword_s"]
        timeout_gap_s = None
        if response_end_s is not None and markers["mode_zero_s"] is not None:
            timeout_gap_s = round(markers["mode_zero_s"] - response_end_s, 3)
        timeout_to_marker_s = None
        if response_end_s is not None and markers["timeout_marker_s"] is not None:
            timeout_to_marker_s = round(markers["timeout_marker_s"] - response_end_s, 3)
        wake_to_mode_zero_s = None
        if markers["wake_frame_s"] is not None and markers["mode_zero_s"] is not None:
            wake_to_mode_zero_s = round(markers["mode_zero_s"] - markers["wake_frame_s"], 3)
        wake_to_timeout_s = None
        if markers["wake_frame_s"] is not None and markers["timeout_marker_s"] is not None:
            wake_to_timeout_s = round(markers["timeout_marker_s"] - markers["wake_frame_s"], 3)
        evidence = self.write_step_artifacts(
            name,
            start_offsets,
            {
                "mode": "wake_timeout_probe",
                "wake_text": wake_text,
                "audio_file": str(audio_path),
                "cached": cached,
                "playback_output": playback_output,
                "wait_s": wait_s,
                "markers": markers,
                "timeout_gap_s": timeout_gap_s,
                "timeout_from_response_end_s": timeout_gap_s,
                "timeout_from_response_end_to_timeout_marker_s": timeout_to_marker_s,
                "wake_to_mode_zero_s": wake_to_mode_zero_s,
                "wake_to_timeout_s": wake_to_timeout_s,
            },
        )
        self.log_event("wake_timeout_probe_finish", {"name": name, "timeout_gap_s": timeout_gap_s, "log_bytes": evidence.log_bytes, "proto_bytes": evidence.proto_bytes})
        return evidence

    def run_post_command_timeout_probe(self, name: str, wake_text: str, command_text: str, wait_s: float = 24.0) -> StepEvidence:
        assert self.log_port and self.proto_port
        self.log_event("post_command_timeout_probe_start", {"name": name, "wake_text": wake_text, "command_text": command_text, "wait_s": wait_s})
        start_offsets = self.slice_offsets()
        probe_started_at = time.time()
        local_log = bytearray()
        local_proto = bytearray()
        markers = {
            "wake_frame_s": None,
            "command_frame_s": None,
            "last_play_start_s": None,
            "last_play_stop_s": None,
            "last_pre_timeout_play_stop_s": None,
            "timeout_marker_s": None,
            "mode_zero_s": None,
        }
        counts = {"play_start": 0, "play_stop": 0}
        audio_items: list[dict[str, Any]] = []

        def update_markers(now: float) -> None:
            text = decode_text(local_log)
            elapsed = round(now - probe_started_at, 3)
            frames = proto_frames_from_hex(local_proto.hex(" ").upper())
            if markers["wake_frame_s"] is None and "A5 FA 01 BB" in frames:
                markers["wake_frame_s"] = elapsed
            if markers["command_frame_s"] is None and "A5 FA 04 BB" in frames:
                markers["command_frame_s"] = elapsed
            play_start_count = text.count("play start")
            if play_start_count > counts["play_start"]:
                counts["play_start"] = play_start_count
                markers["last_play_start_s"] = elapsed
            play_stop_count = text.count("play stop")
            if play_stop_count > counts["play_stop"]:
                counts["play_stop"] = play_stop_count
                markers["last_play_stop_s"] = elapsed
                if markers["timeout_marker_s"] is None:
                    markers["last_pre_timeout_play_stop_s"] = elapsed
            if markers["timeout_marker_s"] is None and "TIME_OUT" in text:
                markers["timeout_marker_s"] = elapsed
            if markers["mode_zero_s"] is None and "MODE=0" in text:
                markers["mode_zero_s"] = elapsed

        def play_text(text: str, label_suffix: str) -> None:
            audio_path, cached = ensure_cached_tts(text=text, voice="Microsoft Huihui Desktop", rate=0, label=f"{sanitize_name(name)}_{label_suffix}")
            audio_items.append({"text": text, "audio_file": str(audio_path), "cached": cached})
            cmd = [sys.executable, str(self.listenai_play), "play", "--audio-file", str(audio_path), "--device-key", self.device_key]
            proc = subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            while True:
                now = time.time()
                log_data = self.log_port.read(4096)
                if log_data:
                    self.log_chunks.extend(log_data)
                    local_log.extend(log_data)
                    update_markers(now)
                proto_data = self.proto_port.read(4096)
                if proto_data:
                    self.proto_chunks.extend(proto_data)
                    local_proto.extend(proto_data)
                    update_markers(now)
                if proc.poll() is not None:
                    break
                time.sleep(0.02)

        play_text(wake_text, "wake")
        settle_deadline = time.time() + DEFAULT_BETWEEN_MAX_WAIT_S
        quiet_since: float | None = None
        while time.time() < settle_deadline:
            now = time.time()
            saw_data = False
            log_data = self.log_port.read(4096)
            if log_data:
                self.log_chunks.extend(log_data)
                local_log.extend(log_data)
                update_markers(now)
                saw_data = True
            proto_data = self.proto_port.read(4096)
            if proto_data:
                self.proto_chunks.extend(proto_data)
                local_proto.extend(proto_data)
                update_markers(now)
                saw_data = True
            if saw_data:
                quiet_since = now
            elif quiet_since is None:
                quiet_since = now
            if markers["last_play_stop_s"] is not None and quiet_since is not None and (now - quiet_since) >= DEFAULT_QUIET_WINDOW_S:
                break
            time.sleep(0.02)
        play_text(command_text, "command")
        while time.time() - probe_started_at < wait_s:
            now = time.time()
            log_data = self.log_port.read(4096)
            if log_data:
                self.log_chunks.extend(log_data)
                local_log.extend(log_data)
                update_markers(now)
            proto_data = self.proto_port.read(4096)
            if proto_data:
                self.proto_chunks.extend(proto_data)
                local_proto.extend(proto_data)
                update_markers(now)
            if markers["mode_zero_s"] is not None:
                break
            time.sleep(0.02)

        timeout_from_response_end_s = None
        response_end_s = markers["last_pre_timeout_play_stop_s"] or markers["last_play_stop_s"]
        if response_end_s is not None and markers["mode_zero_s"] is not None:
            timeout_from_response_end_s = round(markers["mode_zero_s"] - response_end_s, 3)
        timeout_from_response_end_to_timeout_marker_s = None
        if response_end_s is not None and markers["timeout_marker_s"] is not None:
            timeout_from_response_end_to_timeout_marker_s = round(markers["timeout_marker_s"] - response_end_s, 3)
        evidence = self.write_step_artifacts(
            name,
            start_offsets,
            {
                "mode": "post_command_timeout_probe",
                "wake_text": wake_text,
                "command_text": command_text,
                "wait_s": wait_s,
                "markers": markers,
                "response_end_s": response_end_s,
                "timeout_from_response_end_s": timeout_from_response_end_s,
                "timeout_from_response_end_to_timeout_marker_s": timeout_from_response_end_to_timeout_marker_s,
                "audio_items": audio_items,
            },
        )
        self.log_event("post_command_timeout_probe_finish", {"name": name, "timeout_from_response_end_s": timeout_from_response_end_s, "log_bytes": evidence.log_bytes, "proto_bytes": evidence.proto_bytes})
        return evidence

    def run_session_timeout_trial(
        self,
        name: str,
        wake_text: str,
        command_text: str,
        delay_s: float,
        wake_frame: str,
        command_frame: str,
        post_wait_s: float = 3.0,
    ) -> StepEvidence:
        assert self.log_port and self.proto_port
        self.log_event(
            "session_timeout_trial_start",
            {
                "name": name,
                "wake_text": wake_text,
                "command_text": command_text,
                "delay_s": delay_s,
                "wake_frame": wake_frame,
                "command_frame": command_frame,
            },
        )
        start_offsets = self.slice_offsets()
        trial_started_at = time.time()
        local_proto = bytearray()
        audio_items: list[dict[str, Any]] = []
        playback_output: list[dict[str, Any]] = []
        wake_frame_at: float | None = None

        def record_serial(now: float) -> None:
            nonlocal wake_frame_at
            log_data = self.log_port.read(4096)
            if log_data:
                self.log_chunks.extend(log_data)
            proto_data = self.proto_port.read(4096)
            if proto_data:
                self.proto_chunks.extend(proto_data)
                local_proto.extend(proto_data)
                if wake_frame_at is None:
                    frames = proto_frames_from_hex(local_proto.hex(" ").upper())
                    if wake_frame in frames:
                        wake_frame_at = now

        def play_text(text: str, label_suffix: str) -> list[str]:
            audio_path, cached = ensure_cached_tts(text=text, voice="Microsoft Huihui Desktop", rate=0, label=f"{sanitize_name(name)}_{label_suffix}")
            audio_items.append({"text": text, "audio_file": str(audio_path), "cached": cached})
            cmd = [sys.executable, str(self.listenai_play), "play", "--audio-file", str(audio_path), "--device-key", self.device_key]
            proc = subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            local_output: list[str] = []
            while True:
                now = time.time()
                record_serial(now)
                if proc.poll() is not None:
                    break
                time.sleep(0.02)
            playback_output.append({"text": text, "output": local_output})
            return local_output

        play_text(wake_text, "wake")
        reference_at = wake_frame_at if wake_frame_at is not None else time.time()
        while time.time() < reference_at + delay_s:
            record_serial(time.time())
            time.sleep(0.02)
        play_text(command_text, "command")
        self.pump(post_wait_s)

        evidence = self.write_step_artifacts(
            name,
            start_offsets,
            {
                "mode": "session_timeout_trial",
                "wake_text": wake_text,
                "command_text": command_text,
                "delay_s": delay_s,
                "wake_frame": wake_frame,
                "command_frame": command_frame,
                "wake_frame_at_s": None if wake_frame_at is None else round(wake_frame_at - trial_started_at, 3),
                "playback_output": playback_output,
                "audio_items": audio_items,
                "post_wait_s": post_wait_s,
            },
        )
        frames = proto_frames_from_hex(evidence.proto_hex)
        command_ok = command_frame in frames
        evidence.detail["frames"] = frames
        evidence.detail["command_ok"] = command_ok
        evidence.detail["measured_command_delay_s"] = delay_s
        (evidence.step_dir / "meta.json").write_text(json.dumps(evidence.detail, ensure_ascii=False, indent=2), encoding="utf-8")
        self.log_event(
            "session_timeout_trial_finish",
            {
                "name": name,
                "delay_s": delay_s,
                "command_ok": command_ok,
                "log_bytes": evidence.log_bytes,
                "proto_bytes": evidence.proto_bytes,
            },
        )
        return evidence

    def run_powercycle_step(self, name: str, commands: list[str] | None = None, cmd_delay_s: float = 0.35, capture_s: float = 10.0, ready_wait_s: float = 8.0) -> StepEvidence:
        if commands is None:
            commands = ["uut-switch1.off", "uut-switch2.off", "uut-switch1.on"]
        self.log_event("powercycle_start", {"name": name, "commands": commands})
        start_offsets = self.slice_offsets()
        ctrl_port = serial.Serial(CTRL_PORT, baudrate=CTRL_BAUD, timeout=0.05, write_timeout=0.5)
        try:
            ctrl_port.reset_input_buffer()
            for item in commands:
                ctrl_port.write((item + "\r\n").encode("ascii"))
                ctrl_port.flush()
                time.sleep(cmd_delay_s)
            self.pump(capture_s)
            if ready_wait_s > 0:
                self.pump(ready_wait_s)
        finally:
            ctrl_port.close()
        self.ensure_runtime_loglevel(capture_s=1.0, clear_buffers=False)
        evidence = self.write_step_artifacts(name, start_offsets, {"mode": "powercycle", "commands": commands, "cmd_delay_s": cmd_delay_s, "capture_s": capture_s, "ready_wait_s": ready_wait_s})
        self.log_event("powercycle_finish", {"name": name, "log_bytes": evidence.log_bytes, "proto_bytes": evidence.proto_bytes})
        return evidence

    def save_streams(self) -> None:
        (self.stream_dir / "com38_raw.bin").write_bytes(bytes(self.log_chunks))
        com38_text = decode_text(bytes(self.log_chunks))
        (self.stream_dir / "com38_utf8.txt").write_text(com38_text, encoding="utf-8")
        (self.stream_dir / "com38.log").write_text(com38_text, encoding="utf-8")
        (self.stream_dir / "com36_raw.bin").write_bytes(bytes(self.proto_chunks))
        com36_hex = bytes(self.proto_chunks).hex(" ").upper()
        (self.stream_dir / "com36_hex.txt").write_text(com36_hex, encoding="utf-8")
        (self.stream_dir / "com36.log").write_text(format_proto_log(bytes(self.proto_chunks)), encoding="utf-8")
        meta = {
            "log_port": f"{LOG_PORT}@{LOG_BAUD}",
            "proto_port": f"{PROTO_PORT}@{PROTO_BAUD}",
            "ctrl_port": f"{CTRL_PORT}@{CTRL_BAUD}",
            "device_key": self.device_key,
            "events": len(self.events),
            "saved_at": iso_now(),
            "log_bytes": len(self.log_chunks),
            "proto_bytes": len(self.proto_chunks),
        }
        (self.stream_dir / "stream_meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    def write_case_results(self) -> None:
        payload = {"generated_at": iso_now(), "bundle_dir": str(self.bundle_dir.relative_to(ROOT)), "case_results": self.case_results}
        self.case_results_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        lines = [
            "# 本轮目录重构后全链路执行结果",
            "",
            f"- 交付目录：`{self.bundle_dir.relative_to(ROOT)}`",
            f"- 烧录日志目录：`{self.burn_dir.relative_to(ROOT)}`",
            f"- 连续串口日志目录：`{self.stream_dir.relative_to(ROOT)}`",
            f"- 用例结果 JSON：`{self.case_results_path.relative_to(ROOT)}`",
            "",
            "## 用例结果",
            "",
            "| 用例ID | 模块 | 状态 | 结论 | 证据 |",
            "| --- | --- | --- | --- | --- |",
        ]
        for item in self.case_results:
            evidence = "<br>".join(f"`{path}`" for path in item["evidence"])
            lines.append(f"| `{item['case_id']}` | {item['module']} | `{item['status']}` | {item['summary']} | {evidence} |")
        self.summary_md_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    def write_testability_gate(self, payload: dict[str, Any]) -> None:
        self.testability_gate = payload
        self.testability_gate_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def write_failure_analysis(self) -> None:
        requirements = self.spec["requirements"]
        volume_persist_expected = is_yes(requirements["volume_power_save_raw"])
        auto_pass = sum(1 for item in self.case_results if item["status"] == "PASS")
        auto_fail = [item for item in self.case_results if item["status"] == "FAIL"]
        auto_todo = [item for item in self.case_results if item["status"] == "TODO"]
        gate = self.testability_gate or {}
        if gate and not gate.get("passed", True):
            lines = [
                "# Failure Analysis",
                "",
                "## 可测性门禁未通过",
                "",
                "- 当前固件未满足“可进入需求测试”的前置条件，本轮在门禁失败后立即停止后续用例执行。",
                f"- 首次上电默认音量：`{gate.get('first_boot_config', {}).get('volume', 'missing')}`",
                f"- 首启 Running Config 次数：`{gate.get('startup_running_config_count', 0)}`",
                f"- 首启 RESET 次数：`{gate.get('startup_reset_count', 0)}`",
                f"- 首启后 6s 观察窗口 Running Config 次数：`{gate.get('idle_running_config_count', 0)}`",
                f"- 首启后 6s 观察窗口 RESET 次数：`{gate.get('idle_reset_count', 0)}`",
                f"- 门禁阶段算法异常次数：`{gate.get('algo_fail_count', 0)}`",
                f"- 默认唤醒词+普通命令协议链路：`{gate.get('interaction_frames', [])}`",
                "",
                "## 阻断原因",
                "",
            ]
            for reason in gate.get("reasons", []) or ["门禁条件未满足，但缺少详细原因。"]:
                lines.append(f"- {reason}")
            lines.extend(
                [
                    "",
                    "## 结论",
                    "",
                    "- 当前固件不具备可测试状态，后续需求项结果不应继续判定 PASS / FAIL。",
                    "- 建议先修复重复重启、算法初始化失败或默认唤醒链路异常，再重新进入完整全链路测试。",
                ]
            )
            self.failure_analysis_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
            return
        fail_map = {item["case_id"]: item for item in auto_fail}
        fail_list_line = "- 本轮无 FAIL"
        if auto_fail:
            fail_labels = ", ".join(f"`{item['case_id']}`" for item in auto_fail)
            fail_list_line = f"- 当前 FAIL 列表：{fail_labels}"
        lines = [
            "# Failure Analysis",
            "",
            "## 本轮执行概览",
            "",
            f"- 本轮自动/半自动汇总结果：`PASS={auto_pass}`、`FAIL={len(auto_fail)}`、`TODO={len(auto_todo)}`",
            f"- 人工保留项：`{auto_todo[0]['case_id']}`（上电欢迎语）" if auto_todo else "- 本轮无人工保留项",
            fail_list_line,
            "",
            "## 需求一致性结论",
            "",
            "- 本轮已把“功能是否正常”“参数是否一致”“异常场景是否稳定”拆开判定，避免把功能可用误判成需求完全通过",
            f"- 协议判断以 `{PROTO_PORT}` 原始抓取或发送 payload 为准，`{LOG_PORT}` 仅用于接收 / 播报 / 状态辅助",
            "- 语音注册的学习次数、失败重试上限、模板数本轮已单列为配置一致性项，独立给出结论",
            "",
            "## FAIL 归因",
            "",
        ]

        if "CFG-WAKE-001" in fail_map:
            item = fail_map["CFG-WAKE-001"]
            lines.extend(
                [
                    "### `CFG-WAKE-001`",
                    "",
                    "- 模块：配置一致性-会话参数",
                    f"- 当前现象：实测唤醒会话时长约 `{item['detail'].get('timeout_gap_s')}`s，需求为 `{requirements['wake_timeout_s']}s`",
                    "- 影响：功能上仍能超时退出，但超时数值与需求不一致，不能按 PASS 结论关闭",
                    "- 修复建议：核对会话超时参数入包值和运行态实际超时逻辑，确认没有沿用其他旧配置",
                    "",
                ]
            )

        if "CFG-VOL-001" in fail_map:
            item = fail_map["CFG-VOL-001"]
            boot_cfg = item["detail"].get("boot_config", {})
            lines.extend(
                [
                    "### `CFG-VOL-001`",
                    "",
                    "- 模块：配置一致性-音量参数",
                    f"- 当前现象：冷启动默认音量为 `{boot_cfg.get('volume')}`，需求为 `{requirements['default_volume']}`",
                    "- 影响：音量功能本身可用，但默认值不满足需求",
                    "- 修复建议：检查默认配置表或 `config.clear` 后的默认音量初始化值",
                    "",
                ]
            )

        if "CFG-VOL-002" in fail_map:
            item = fail_map["CFG-VOL-002"]
            lines.extend(
                [
                    "### `CFG-VOL-002`",
                    "",
                    "- 模块：配置一致性-音量参数",
                    f"- 当前现象：实测可达音量档位为 `{item['detail'].get('values')}`，不是需求的 `{requirements['volume_steps']}` 档",
                    "- 影响：音量调节功能可用，但档位数与需求不一致",
                    "- 修复建议：检查音量等级映射和边界值定义，确认运行态档位数与需求表一致",
                    "",
                ]
            )

        if "VOL-003" in fail_map:
            item = fail_map["VOL-003"]
            boot_cfg = item["detail"].get("boot_config", {})
            lines.extend(
                [
                    "### `VOL-003`",
                    "",
                    "- 模块：音量控制",
                    (
                        f"- 当前现象：断电前目标音量为 `{item['detail'].get('target_volume', 'missing')}`，重启后 `volume` 为 `{boot_cfg.get('volume')}`，未保持断电前档位"
                        if volume_persist_expected
                        else f"- 当前现象：将音量设到非默认档位后断电，重启 `volume` 仍为 `{boot_cfg.get('volume')}`，未恢复默认需求值 `{requirements['default_volume']}`"
                    ),
                    (
                        "- 影响：音量掉电保持需求不成立"
                        if volume_persist_expected
                        else "- 影响：音量掉电不保持需求不成立"
                    ),
                    (
                        "- 修复建议：检查音量配置的持久化写入和上电恢复逻辑，确认保存成功后能正确回读"
                        if volume_persist_expected
                        else "- 修复建议：检查音量配置的保存位开关，确认音量是否被错误写入持久化配置"
                    ),
                    "",
                ]
            )

        if "REG-CONFLICT-001" in fail_map:
            lines.extend(
                [
                    "### `REG-CONFLICT-001`",
                    "",
                    "- 模块：语音注册-冲突词",
                    "- 当前现象：功能词 `增大音量` 被错误学习成自定义命令词，回测时触发成了 `打开电风扇`，不是原本的音量增大动作",
                    "- 问题定位：命令词学习态下，对“功能词 / 保留词 / 可学习词”的边界校验不完整，导致功能词样本被错误写入学习结果",
                    "- 修复建议：",
                    "  1. 在命令词学习入口增加功能词黑名单校验，拒绝把已有功能控制词写入学习结果",
                    "  2. 学习成功前增加目标协议冲突检查，避免把功能协议错误映射到其他动作",
                    "  3. 回归补测 `增大音量`、`最小音量`、`关闭播报`、`退出识别` 等功能词冲突集",
                    "",
                ]
            )

        other_fail_ids = [case_id for case_id in fail_map if case_id not in {"CFG-WAKE-001", "CFG-VOL-001", "CFG-VOL-002", "VOL-003", "REG-CONFLICT-001"}]
        for case_id in other_fail_ids:
            item = fail_map[case_id]
            lines.extend(
                [
                    f"### `{case_id}`",
                    "",
                    f"- 模块：{item['module']}",
                    f"- 当前现象：{item['summary']}",
                    f"- 证据：`{item['evidence'][0]}`" if item["evidence"] else "- 证据：见 `case_results.json`",
                    "",
                ]
            )

        lines.extend(
            [
                "## 当前保留的人工项",
                "",
                "### `SESS-001`",
                "",
                "- 原因：上电欢迎语更适合人工听感 / 现场观察校验",
                "- 本轮自动辅助证据：",
                "  - `01_assist_startup_powercycle_capture`",
                "  - `streams/com38_utf8.txt`",
                "- 结论：自动链路已补足启动连续日志，但最终通过仍以人工确认欢迎语和待机就绪为准",
            ]
        )
        self.failure_analysis_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    def sync_bundle_root_artifacts(self) -> None:
        sync_map = {
            "测试方案.md": PLAN_PATH,
            "测试用例-正式版.xlsx": CASE_XLSX_PATH,
            "case_results.json": self.case_results_path,
            "execution_summary.md": self.summary_md_path,
            "failure_analysis.md": self.failure_analysis_path,
            "testability_gate.json": self.testability_gate_path,
            "burn.log": self.burn_dir / "burn.log",
            "burn_tool.log": self.burn_dir / "burn_tool.log",
            "com38.log": self.stream_dir / "com38.log",
            "com36.log": self.stream_dir / "com36.log",
        }
        for output_name, src in sync_map.items():
            if src.exists():
                shutil.copy2(src, self.bundle_dir / output_name)




def step_has_play_id(evidence: StepEvidence, play_id: int) -> bool:
    return re.search(rf"play id\s*:\s*{play_id}\b", evidence.log_text) is not None


def run_default_volume_position_probe(runner: FullflowRunner, name: str, volume_steps: int) -> dict[str, Any]:
    max_overflow_id = 23
    min_overflow_id = 24
    max_iterations = int(volume_steps) + 3
    evidence: list[StepEvidence] = []
    observed_levels: list[int] = []
    up_success_steps = 0
    last_level: int | None = None

    for index in range(1, max_iterations + 1):
        step = runner.run_voice_sequence(f"{name}_default_to_max_up_{index}", ["小度小度", "大声点"], post_wait_s=4.0)
        evidence.append(step)
        level = last_runtime_volume_level(step.log_text)
        if isinstance(level, int):
            observed_levels.append(level)
        if step_has_play_id(step, max_overflow_id):
            break
        if level is None:
            break
        if last_level is not None and level <= last_level:
            break
        up_success_steps += 1
        last_level = level

    last_level = observed_levels[-1] if observed_levels else None
    for index in range(1, max_iterations + 1):
        step = runner.run_voice_sequence(f"{name}_max_to_min_down_{index}", ["小度小度", "小声点"], post_wait_s=4.0)
        evidence.append(step)
        level = last_runtime_volume_level(step.log_text)
        if isinstance(level, int):
            observed_levels.append(level)
        if step_has_play_id(step, min_overflow_id):
            break
        if level is None:
            break
        if last_level is not None and level >= last_level:
            break
        last_level = level

    last_level = observed_levels[-1] if observed_levels else None
    for index in range(1, max_iterations + 1):
        step = runner.run_voice_sequence(f"{name}_min_to_max_up_{index}", ["小度小度", "大声点"], post_wait_s=4.0)
        evidence.append(step)
        level = last_runtime_volume_level(step.log_text)
        if isinstance(level, int):
            observed_levels.append(level)
        if step_has_play_id(step, max_overflow_id):
            break
        if level is None:
            break
        if last_level is not None and level <= last_level:
            break
        last_level = level

    unique_levels = ordered_unique([level for level in observed_levels if isinstance(level, int)])
    total_levels = len(unique_levels)
    inferred_default_gear = total_levels - up_success_steps if total_levels > 0 else None
    return {
        "evidence": evidence,
        "up_success_steps_to_max": up_success_steps,
        "observed_runtime_levels": observed_levels,
        "unique_runtime_levels": unique_levels,
        "sorted_runtime_levels": sorted(set(unique_levels)),
        "total_levels": total_levels,
        "inferred_default_gear": inferred_default_gear,
    }

def step_pass(evidence: StepEvidence, *, require_proto: bool | None = None, markers: list[str] | None = None) -> bool:
    if require_proto is True and evidence.proto_bytes <= 0:
        return False
    if require_proto is False and evidence.proto_bytes > 0:
        return False
    if markers:
        return all(marker in evidence.log_text for marker in markers)
    return True


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def has_all_markers(text: str, markers: list[str]) -> bool:
    return all(marker in text for marker in markers)


def export_cases() -> None:
    cmd = [sys.executable, str(CASES_DIR / "export_case_md_to_xlsx.py"), "--input", str(CASE_MD_PATH), "--output", str(CASE_XLSX_PATH)]
    subprocess.run(cmd, cwd=ROOT, check=True)


def regenerate_formal_assets() -> None:
    cmd = [sys.executable, str(GENERATE_ASSETS_SCRIPT)]
    subprocess.run(cmd, cwd=ROOT, check=True)


def parse_case_table_line(line: str) -> tuple[str, list[str]] | None:
    if not line.startswith("| `"):
        return None
    cells = [cell.strip() for cell in line[1:-1].split("|")]
    if len(cells) != 10:
        return None
    case_id = cells[0].strip("`")
    return case_id, cells


def collect_statuses_from_lines(lines: list[str]) -> dict[str, str]:
    status_map: dict[str, str] = {}
    for line in lines:
        parsed = parse_case_table_line(line)
        if not parsed:
            continue
        case_id, cells = parsed
        status_map[case_id] = cells[8].strip("`")
    return status_map


def update_case_markdown(case_results: list[dict[str, Any]], evidence_map: dict[str, list[Path]]) -> None:
    status_map = {item["case_id"]: item["status"] for item in case_results}
    lines = CASE_MD_PATH.read_text(encoding="utf-8").splitlines()
    updated: list[str] = []
    for line in lines:
        parsed = parse_case_table_line(line)
        if not parsed:
            updated.append(line)
            continue
        case_id, cells = parsed
        if case_id not in status_map:
            updated.append(line)
            continue
        cells[8] = f"`{status_map[case_id]}`"
        evidence_paths = evidence_map.get(case_id, [])
        if evidence_paths:
            evidence = "<br>".join(f"`{path.relative_to(ROOT)}`" for path in evidence_paths)
            cells[9] = f"{evidence}<br>目录重构后复跑闭环"
        updated.append("| " + " | ".join(cells) + " |")
    final_status_map = collect_statuses_from_lines(updated)
    pass_count = sum(1 for status in final_status_map.values() if status == "PASS")
    fail_ids = [case_id for case_id, status in final_status_map.items() if status == "FAIL"]
    todo_ids = [case_id for case_id, status in final_status_map.items() if status == "TODO"]
    text = "\n".join(updated)
    text = re.sub(r"- 当前状态分布：`PASS=\d+`、`FAIL=\d+`、`TODO=\d+`", f"- 当前状态分布：`PASS={pass_count}`、`FAIL={len(fail_ids)}`、`TODO={len(todo_ids)}`", text)
    fail_line = "- 当前明确保留缺陷："
    fail_line += "无" if not fail_ids else "、".join(f"`{case_id}`" for case_id in fail_ids)
    text = re.sub(r"- 当前(?:唯一)?明确保留缺陷：.*", fail_line, text)
    todo_header = "- 当前仅剩人工保留项 `TODO` 为："
    todo_block = todo_header
    if todo_ids:
        todo_block += "\n" + "\n".join(f"  - `{case_id}`" for case_id in todo_ids)
    else:
        todo_block += "\n  - 无"
    text = re.sub(r"- 当前(?:仅剩人工保留项|优先级最高的) `TODO` 为：\n(?:  - .*\n?)*", todo_block, text)
    CASE_MD_PATH.write_text(text + "\n", encoding="utf-8")

def main() -> int:
    regenerate_formal_assets()
    runner = FullflowRunner()
    evidence_map: dict[str, list[Path]] = {}
    requirements = runner.spec["requirements"]
    words = runner.spec["words"]
    voice_reg = runner.spec["voice_reg"]

    runner.prepare_static_assets()
    runner.burn_firmware()
    runner.open_ports()
    try:
        startup = runner.run_powercycle_step("assist_startup_powercycle_capture")
        startup_cfg = parse_boot_config(startup.log_text)
        burn_log_text = read_text(runner.burn_dir / "burn.log") if (runner.burn_dir / "burn.log").exists() else ""

        env001_status = "PASS" if str(startup_cfg.get("version")) == "1.0.0" else "FAIL"
        runner.add_case_result("ENV-001", "环境确认", env001_status, f"启动配置版本字段为 `{startup_cfg.get('version', 'missing')}`", [startup.step_dir], {"boot_config": startup_cfg})
        evidence_map["ENV-001"] = [startup.step_dir]

        env002_status = "PASS" if "Burn flow completed" in burn_log_text and env001_status == "PASS" else "FAIL"
        runner.add_case_result("ENV-002", "烧录确认", env002_status, "烧录流程成功且启动版本与目标固件一致", [runner.burn_dir / "burn.log", startup.step_dir])
        evidence_map["ENV-002"] = [runner.burn_dir / "burn.log", startup.step_dir]

        runner.add_case_result("SESS-001", "启动待机", "TODO", "本轮保留为人工验证；已补充启动连续日志作为辅助证据", [startup.step_dir])
        evidence_map["SESS-001"] = [startup.step_dir]

        startup_gain = parse_mic_gain(startup.log_text)
        startup_idle_gate = runner.run_idle_wait_step("assist_startup_gate_idle_watch", duration_s=6.0)
        gate_interaction = runner.run_voice_sequence("assist_gate_default_wake_open", ["小度小度", "打开电风扇"], post_wait_s=3.0)
        gate_expected_frames = [words["小度小度"]["发送协议"], words["打开电风扇"]["发送协议"]]
        gate_frames = proto_frames_from_hex(gate_interaction.proto_hex)
        gate_proto_ok = words["打开电风扇"]["发送协议"] in gate_frames
        gate_startup_running_cfg_count = count_occurrences(startup.log_text, "Running Config")
        gate_startup_reset_count = count_occurrences(startup.log_text, "RESET=")
        gate_idle_running_cfg_count = count_occurrences(startup_idle_gate.log_text, "Running Config")
        gate_idle_reset_count = count_occurrences(startup_idle_gate.log_text, "RESET=")
        gate_algo_fail_count = sum(
            count_occurrences(text, marker)
            for text in [startup.log_text, startup_idle_gate.log_text]
            for marker in ["wIvwCreate fail", "ai_create failed", "algo tick same"]
        )
        gate_reasons: list[str] = []
        if gate_startup_running_cfg_count != 1:
            gate_reasons.append(f"首启日志中的 Running Config 次数={gate_startup_running_cfg_count}，不是单次稳定启动")
        if gate_startup_reset_count > 1:
            gate_reasons.append(f"首启日志中的 RESET 次数={gate_startup_reset_count}，存在重复重启迹象")
        if gate_idle_running_cfg_count > 0 or gate_idle_reset_count > 0:
            gate_reasons.append(
                f"首启后 6s 观察窗口内再次出现启动迹象，Running Config={gate_idle_running_cfg_count}，RESET={gate_idle_reset_count}"
            )
        if gate_algo_fail_count > 0:
            gate_reasons.append(f"门禁阶段捕获到算法异常日志，共 {gate_algo_fail_count} 次")
        if not gate_proto_ok:
            gate_reasons.append(f"默认唤醒词后普通命令未形成有效控制闭环，实测帧={gate_frames}")
        gate_passed = not gate_reasons
        gate_payload = {
            "required": True,
            "checked_at": iso_now(),
            "passed": gate_passed,
            "first_boot_config": startup_cfg,
            "first_boot_gain": startup_gain,
            "startup_running_config_count": gate_startup_running_cfg_count,
            "startup_reset_count": gate_startup_reset_count,
            "idle_running_config_count": gate_idle_running_cfg_count,
            "idle_reset_count": gate_idle_reset_count,
            "algo_fail_count": gate_algo_fail_count,
            "interaction_frames": gate_frames,
            "expected_frames": gate_expected_frames,
            "evidence": [
                str(startup.step_dir.relative_to(ROOT)),
                str(startup_idle_gate.step_dir.relative_to(ROOT)),
                str(gate_interaction.step_dir.relative_to(ROOT)),
            ],
            "reasons": gate_reasons,
        }
        runner.write_testability_gate(gate_payload)
        gate_summary = (
            "设备满足可测性门禁：首启稳定、默认唤醒词可唤醒，普通命令交互闭环正常"
            if gate_passed
            else "设备不满足可测性门禁：" + "；".join(gate_reasons)
        )
        runner.add_case_result(
            "ENV-GATE-001",
            "环境确认",
            "PASS" if gate_passed else "FAIL",
            gate_summary,
            [startup.step_dir, startup_idle_gate.step_dir, gate_interaction.step_dir],
            gate_payload,
        )
        evidence_map["ENV-GATE-001"] = [startup.step_dir, startup_idle_gate.step_dir, gate_interaction.step_dir]
        if not gate_passed:
            raise UntestableFirmware(gate_summary)

        default_volume_probe = run_default_volume_position_probe(runner, "cfg_default_volume_position_probe", requirements["volume_steps"])
        default_volume_probe_evidence = [item.step_dir for item in default_volume_probe["evidence"]]
        runner.run_shell_step("assist_clear_after_default_volume_probe", "config.clear", capture_s=3.0, ready_wait_s=1.0)
        runner.run_shell_step("assist_reboot_after_default_volume_probe", "reboot", capture_s=10.0, ready_wait_s=8.0)

        reg_cmd_full_entry = runner.run_voice_sequence("assist_reg_cmd_template_full_entry", ["小度小度", "学习命令词"], post_wait_s=3.0)

        runner.run_shell_step("assist_config_clear_after_burn", "config.clear", capture_s=3.0, ready_wait_s=1.0)
        reboot_after_clear = runner.run_shell_step("assist_reboot_after_config_clear", "reboot", capture_s=10.0, ready_wait_s=8.0)
        clean_cfg = parse_boot_config(reboot_after_clear.log_text)
        cfg_audio_summary = (
            f"首启日志可见 mic 增益片段={startup_gain.get('analog_gain_db', 'missing')}/{startup_gain.get('digital_gain_db', 'missing')}dB，"
            f"需求={requirements['mic_analog_gain_db']}/{requirements['mic_digital_gain_db']}dB；当前保留人工确认"
        )
        runner.add_case_result(
            "CFG-AUDIO-001",
            "配置一致性-基础参数",
            "TODO",
            cfg_audio_summary,
            [startup.step_dir],
            {"boot_config": startup_cfg, "mic_gain": startup_gain, "manual_required": True},
        )
        evidence_map["CFG-AUDIO-001"] = [startup.step_dir]

        default_volume_inferred = default_volume_probe.get("inferred_default_gear")
        expected_raw_zero_based = requirements["default_volume"] - 1
        if not isinstance(default_volume_inferred, int):
            cfg_default_volume_status = "BLOCKED"
            cfg_default_volume_summary = "默认音量探测未形成完整档位证据，不能判定默认值"
        else:
            cfg_default_volume_status = "PASS" if default_volume_inferred == requirements["default_volume"] else "FAIL"
            cfg_default_volume_summary = (
                f"烧录后探测默认音量档位={default_volume_inferred}，需求={requirements['default_volume']}；"
                f"启动配置 raw volume={startup_cfg.get('volume', 'missing')}，期望 raw≈{expected_raw_zero_based}"
            )
        cfg_default_volume_detail = {
            "boot_config": startup_cfg,
            "first_boot_required": True,
            "expected_default_volume": requirements["default_volume"],
            "expected_raw_zero_based": expected_raw_zero_based,
            "raw_volume_from_boot": startup_cfg.get("volume"),
            "probe": {key: value for key, value in default_volume_probe.items() if key != "evidence"},
        }
        runner.add_case_result(
            "CFG-VOL-001",
            "配置一致性-音量参数",
            cfg_default_volume_status,
            cfg_default_volume_summary,
            [startup.step_dir, *default_volume_probe_evidence],
            cfg_default_volume_detail,
        )
        evidence_map["CFG-VOL-001"] = [startup.step_dir, *default_volume_probe_evidence]

        wake_expected = float(requirements["wake_timeout_s"])
        timeout_probe = runner.run_wake_timeout_probe("cfg_wake_timeout_probe", "小度小度", wait_s=float(requirements["wake_timeout_s"] + 18))
        timeout_probe_settle = runner.run_idle_wait_step("cfg_wake_timeout_probe_settle", duration_s=3.0)
        post_cmd_timeout_probe = runner.run_post_command_timeout_probe(
            "cfg_wake_timeout_post_command_probe",
            "小度小度",
            "打开电风扇",
            wait_s=float(requirements["wake_timeout_s"] + 30),
        )
        measured_timeout = timeout_probe.detail.get("timeout_from_response_end_s")
        measured_timeout_to_marker = timeout_probe.detail.get("timeout_from_response_end_to_timeout_marker_s")
        wake_to_timeout = timeout_probe.detail.get("wake_to_timeout_s")
        wake_to_mode_zero = timeout_probe.detail.get("wake_to_mode_zero_s")
        post_command_timeout = post_cmd_timeout_probe.detail.get("timeout_from_response_end_s")
        post_command_timeout_to_marker = post_cmd_timeout_probe.detail.get("timeout_from_response_end_to_timeout_marker_s")
        timeout_delta = None
        if isinstance(measured_timeout, (int, float)) and isinstance(post_command_timeout, (int, float)):
            timeout_delta = round(abs(measured_timeout - post_command_timeout), 3)
        cfg_wake_status = (
            "PASS"
            if isinstance(measured_timeout, (int, float))
            and isinstance(post_command_timeout, (int, float))
            and abs(measured_timeout - wake_expected) <= 1.5
            and abs(post_command_timeout - wake_expected) <= 1.5
            and (timeout_delta is not None and timeout_delta <= 1.0)
            else "FAIL"
        )
        timeout_detail = {
            "mode": "wake_timeout_marker_probe",
            "expected_timeout_s": wake_expected,
            "measured_timeout_s": measured_timeout,
            "timeout_from_response_end_s": measured_timeout,
            "timeout_from_response_end_to_timeout_marker_s": measured_timeout_to_marker,
            "wake_to_timeout_s": wake_to_timeout,
            "wake_to_mode_zero_s": wake_to_mode_zero,
            "measured_upper_bound_s": measured_timeout,
            "post_command_timeout_from_response_end_s": post_command_timeout,
            "post_command_timeout_from_response_end_to_timeout_marker_s": post_command_timeout_to_marker,
            "timeout_delta_s": timeout_delta,
            "pure_wake_markers": timeout_probe.detail.get("markers", {}),
            "post_command_markers": post_cmd_timeout_probe.detail.get("markers", {}),
            "evidence": [
                str(timeout_probe.step_dir.relative_to(ROOT)),
                str(timeout_probe_settle.step_dir.relative_to(ROOT)),
                str(post_cmd_timeout_probe.step_dir.relative_to(ROOT)),
            ],
        }
        runner.add_case_result(
            "CFG-WAKE-001",
            "配置一致性-会话参数",
            cfg_wake_status,
            (
                f"纯唤醒响应结束 -> 超时 `{measured_timeout}`s，"
                f"命令响应结束 -> 超时 `{post_command_timeout}`s，"
                f"两场景差值 `{timeout_delta}`s，需求 `{requirements['wake_timeout_s']}s`"
            ),
            [timeout_probe.step_dir, timeout_probe_settle.step_dir, post_cmd_timeout_probe.step_dir],
            timeout_detail,
        )
        evidence_map["CFG-WAKE-001"] = [timeout_probe.step_dir, timeout_probe_settle.step_dir, post_cmd_timeout_probe.step_dir]

        passive_report = runner.run_protocol_step("cfg_passive_report_proto", words["播报语"]["接收协议"], post_wait_s=4.0, pre_wait_s=0.3)
        passive_report_status = "PASS" if passive_report.detail.get("payload_hex") == words["播报语"]["接收协议"] and has_all_markers(passive_report.log_text, ["receive msg:: A5 FB 12 CC", "play start", "play id : 18", "play stop"]) else "FAIL"
        runner.add_case_result(
            "CFG-PROTO-003",
            "配置一致性-协议",
            passive_report_status,
            f"被动播报协议发送 `{passive_report.detail.get('payload_hex')}`，日志已{'命中' if passive_report_status == 'PASS' else '未完整命中'}接收与播报链路",
            [passive_report.step_dir],
            passive_report.detail,
        )
        evidence_map["CFG-PROTO-003"] = [passive_report.step_dir]

        volume_persist_expected = is_yes(requirements["volume_power_save_raw"])
        volume_level_evidence: list[Path] = []
        volume_level_up_steps: list[StepEvidence] = []
        volume_level_down_steps: list[StepEvidence] = []

        volume_min_anchor = runner.run_voice_sequence("cfg_volume_level_probe_min_anchor", ["小度小度", "最小音量"], post_wait_s=3.0)
        volume_level_evidence.append(volume_min_anchor.step_dir)
        min_level = last_runtime_volume_level(volume_min_anchor.log_text)

        asc_levels_raw: list[int | None] = []
        current_level = min_level
        for index in range(1, requirements["volume_steps"] + 3):
            step = runner.run_voice_sequence(f"cfg_volume_level_probe_up_{index}", ["小度小度", "大声点"], post_wait_s=3.0)
            volume_level_up_steps.append(step)
            volume_level_evidence.append(step.step_dir)
            level = last_runtime_volume_level(step.log_text)
            asc_levels_raw.append(level)
            if level is None:
                break
            if current_level is not None and level <= current_level:
                break
            current_level = level

        volume_max_anchor = runner.run_voice_sequence("cfg_volume_level_probe_max_anchor", ["小度小度", "最大音量"], post_wait_s=3.0)
        volume_level_evidence.append(volume_max_anchor.step_dir)
        max_level = last_runtime_volume_level(volume_max_anchor.log_text)

        desc_levels_raw: list[int | None] = []
        current_level = max_level
        for index in range(1, requirements["volume_steps"] + 3):
            step = runner.run_voice_sequence(f"cfg_volume_level_probe_down_{index}", ["小度小度", "小声点"], post_wait_s=3.0)
            volume_level_down_steps.append(step)
            volume_level_evidence.append(step.step_dir)
            level = last_runtime_volume_level(step.log_text)
            desc_levels_raw.append(level)
            if level is None:
                break
            if current_level is not None and level >= current_level:
                break
            current_level = level

        volume_min_setup = runner.run_voice_sequence("cfg_volume_persist_probe_min", ["小度小度", "最小音量"], post_wait_s=4.0)
        volume_min_save_wait = runner.run_idle_wait_step("cfg_volume_persist_wait_refresh", duration_s=8.0)
        volume_power_boot = runner.run_powercycle_step("cfg_volume_powercycle_reboot", capture_s=10.0, ready_wait_s=8.0)
        volume_power_cfg = parse_boot_config(volume_power_boot.log_text)
        target_volume = 0 if last_runtime_volume_level(volume_min_setup.log_text) is not None else None
        persisted_refresh_values = extract_volume_values(volume_min_setup.log_text + "\n" + volume_min_save_wait.log_text)
        asc_unique_levels = ordered_unique([level for level in [min_level, *asc_levels_raw, max_level] if isinstance(level, int)])
        desc_unique_levels = ordered_unique([level for level in [max_level, *desc_levels_raw] if isinstance(level, int)])
        desc_reversed_levels = list(reversed(desc_unique_levels))
        symmetric_levels = asc_unique_levels == desc_reversed_levels
        cfg_volume_steps_status = (
            "PASS"
            if len(asc_unique_levels) == requirements["volume_steps"]
            and len(desc_unique_levels) == requirements["volume_steps"]
            and symmetric_levels
            else "FAIL"
        )
        runner.add_case_result(
            "CFG-VOL-002",
            "配置一致性-音量参数",
            cfg_volume_steps_status,
            (
                f"最小到最大实测 `{asc_unique_levels}`，"
                f"最大到最小实测 `{desc_unique_levels}`，"
                f"对称性={'一致' if symmetric_levels else '不一致'}，需求={requirements['volume_steps']} 档"
            ),
            [*volume_level_evidence, volume_min_setup.step_dir, volume_min_save_wait.step_dir, volume_power_boot.step_dir],
            {
                "min_level": min_level,
                "max_level": max_level,
                "asc_levels_raw": asc_levels_raw,
                "desc_levels_raw": desc_levels_raw,
                "asc_unique_levels": asc_unique_levels,
                "desc_unique_levels": desc_unique_levels,
                "desc_reversed_levels": desc_reversed_levels,
                "symmetric_levels": symmetric_levels,
            },
        )
        evidence_map["CFG-VOL-002"] = [*volume_level_evidence, volume_min_setup.step_dir, volume_min_save_wait.step_dir, volume_power_boot.step_dir]

        if volume_persist_expected and target_volume is None:
            expected_volume_after_power = None
            vol003_status = "BLOCKED"
        else:
            expected_volume_after_power = target_volume if volume_persist_expected else requirements["default_volume"]
            vol003_status = "PASS" if volume_power_cfg.get("volume") == expected_volume_after_power else "FAIL"
        runner.add_case_result(
            "VOL-003",
            "音量控制",
            vol003_status,
            (
                f"最小音量断电后启动值={volume_power_cfg.get('volume', 'missing')}，需求应保持断电前档位 {expected_volume_after_power}"
                if volume_persist_expected
                else f"最小音量断电后启动值={volume_power_cfg.get('volume', 'missing')}，需求应恢复到默认音量 {requirements['default_volume']}"
            ),
            [volume_min_setup.step_dir, volume_min_save_wait.step_dir, volume_power_boot.step_dir],
            {"boot_config": volume_power_cfg, "target_volume": target_volume, "persist_expected": volume_persist_expected, "persisted_refresh_values": persisted_refresh_values},
        )
        evidence_map["VOL-003"] = [volume_min_setup.step_dir, volume_min_save_wait.step_dir, volume_power_boot.step_dir]

        runner.run_shell_step("assist_config_clear_after_volume_cfg", "config.clear", capture_s=3.0, ready_wait_s=1.0)
        runner.run_shell_step("assist_reboot_after_volume_cfg_clear", "reboot", capture_s=10.0, ready_wait_s=8.0)

        sess007 = runner.run_voice_sequence("sess_idle_open_blocked", ["打开电风扇"], post_wait_s=3.0)
        runner.add_case_result("SESS-007", "待机阻断", "PASS" if step_pass(sess007, require_proto=False) else "FAIL", "未唤醒时直说控制词未产生主动控制协议", [sess007.step_dir])
        evidence_map["SESS-007"] = [sess007.step_dir]

        sess006_exit = runner.run_voice_sequence("sess_exit_sequence_only", ["小度小度", "退出识别"], post_wait_s=3.0)
        sess006_recover = runner.run_voice_sequence("sess_exit_rewake_open_recovery", ["小度小度", "打开电风扇"], post_wait_s=3.0)
        runner.add_case_result("SESS-006", "退出识别", "PASS" if step_pass(sess006_recover, require_proto=True) else "FAIL", "退出识别后重新唤醒，基础控制恢复正常", [sess006_exit.step_dir, sess006_recover.step_dir])
        evidence_map["SESS-006"] = [sess006_exit.step_dir, sess006_recover.step_dir]

        vol004 = runner.run_voice_sequence("volume_rapid_stability", ["小度小度", "大声点", "大声点", "小声点"], post_wait_s=3.0)
        runner.add_case_result("VOL-004", "音量控制", "PASS" if step_pass(vol004, require_proto=True) else "FAIL", "连续快速调音过程中协议持续正常输出，未见失控迹象", [vol004.step_dir])
        evidence_map["VOL-004"] = [vol004.step_dir]

        voice_off = runner.run_voice_sequence("assist_voice_off_setup", ["小度小度", "关闭语音"], post_wait_s=3.0)
        invalid_proto = runner.run_protocol_step("voice_invalid_proto", "A5 FB 0B CC", post_wait_s=4.0, pre_wait_s=0.3)
        invalid_block = runner.run_voice_sequence("voice_invalid_proto_then_wake_open_blocked", ["小度小度", "打开电风扇"], post_wait_s=3.0)
        runner.add_case_result("VOICE-005", "语音开关", "PASS" if step_pass(invalid_block, require_proto=False) else "FAIL", "错误协议未误开语音，语音入口仍保持关闭", [voice_off.step_dir, invalid_proto.step_dir, invalid_block.step_dir])
        evidence_map["VOICE-005"] = [voice_off.step_dir, invalid_proto.step_dir, invalid_block.step_dir]

        valid_proto = runner.run_protocol_step("voice_valid_proto_open", "A5 FB 0A CC", post_wait_s=3.0, pre_wait_s=0.5)
        voice003 = runner.run_voice_sequence("voice_on_then_volume_up", ["小度小度", "大声点"], post_wait_s=3.0)
        runner.add_case_result("VOICE-003", "语音开关", "PASS" if step_pass(voice003, require_proto=True) else "FAIL", "协议开语音后，音量调节链路恢复正常", [valid_proto.step_dir, voice003.step_dir])
        evidence_map["VOICE-003"] = [valid_proto.step_dir, voice003.step_dir]

        switch_next = runner.run_voice_sequence("switch_wake_to_next", ["小度小度", "切换唤醒词"], post_wait_s=3.0)
        swake_idle_current_open = runner.run_idle_wait_step("switch_idle_before_current_wake_open", duration_s=22.0)
        swake004_open = runner.run_voice_sequence("switch_current_wake_open", ["小爱同学", "打开电风扇"], post_wait_s=3.0)
        swake_idle_current_volume = runner.run_idle_wait_step("switch_idle_before_current_wake_volume", duration_s=22.0)
        swake004_volume = runner.run_voice_sequence("switch_current_wake_volume", ["小爱同学", "大声点"], post_wait_s=3.0)
        swake_idle_default_open = runner.run_idle_wait_step("switch_idle_before_default_wake_open", duration_s=22.0)
        swake005_open = runner.run_voice_sequence("switch_default_wake_open", ["小度小度", "打开电风扇"], post_wait_s=3.0)
        swake_idle_default_volume = runner.run_idle_wait_step("switch_idle_before_default_wake_volume", duration_s=22.0)
        swake005_volume = runner.run_voice_sequence("switch_default_wake_volume", ["小度小度", "大声点"], post_wait_s=3.0)
        swake_idle = runner.run_idle_wait_step("switch_idle_before_other_wake_blocked", duration_s=22.0)
        swake003 = runner.run_voice_sequence("switch_other_wake_blocked", ["天猫精灵", "打开电风扇"], post_wait_s=3.0)
        switch_to_tmall = runner.run_voice_sequence("switch_cycle_to_tmall", ["小度小度", "切换唤醒词"], post_wait_s=3.0)
        switch_idle_tmall_open = runner.run_idle_wait_step("switch_idle_before_tmall_open", duration_s=22.0)
        switch_tmall_open = runner.run_voice_sequence("switch_cycle_tmall_open", ["天猫精灵", "打开电风扇"], post_wait_s=3.0)
        switch_back_default = runner.run_voice_sequence("switch_cycle_back_default", ["小度小度", "切换唤醒词"], post_wait_s=3.0)

        runner.add_case_result("SWAKE-003", "切换唤醒词", "PASS" if step_pass(swake003, require_proto=False) else "FAIL", "非当前非默认唤醒词未误唤醒", [switch_next.step_dir, swake_idle.step_dir, swake003.step_dir])
        evidence_map["SWAKE-003"] = [switch_next.step_dir, swake_idle.step_dir, swake003.step_dir]

        runner.add_case_result("SWAKE-004", "切换唤醒词", "PASS" if step_pass(swake004_open, require_proto=True) and step_pass(swake004_volume, require_proto=True) else "FAIL", "切换后的当前唤醒词下，基础控制和音量都正常", [switch_next.step_dir, swake_idle_current_open.step_dir, swake004_open.step_dir, swake_idle_current_volume.step_dir, swake004_volume.step_dir])
        evidence_map["SWAKE-004"] = [switch_next.step_dir, swake_idle_current_open.step_dir, swake004_open.step_dir, swake_idle_current_volume.step_dir, swake004_volume.step_dir]

        swake005_ok = (
            evidence_has_frames(swake005_open, [words["小度小度"]["发送协议"], words["打开电风扇"]["发送协议"]])
            and evidence_has_frame(swake005_volume, words["大声点"]["发送协议"])
        )
        runner.add_case_result("SWAKE-005", "切换唤醒词", "PASS" if swake005_ok else "FAIL", "切换后默认唤醒词仍可完成基础控制和音量交互", [switch_next.step_dir, swake_idle_default_open.step_dir, swake005_open.step_dir, swake_idle_default_volume.step_dir, swake005_volume.step_dir], {"open_frames": proto_frames_from_hex(swake005_open.proto_hex), "volume_frames": proto_frames_from_hex(swake005_volume.proto_hex)})
        evidence_map["SWAKE-005"] = [switch_next.step_dir, swake_idle_default_open.step_dir, swake005_open.step_dir, swake_idle_default_volume.step_dir, swake005_volume.step_dir]

        runner.add_case_result("SWAKE-007", "切换唤醒词", "PASS" if step_pass(switch_tmall_open, require_proto=True) else "FAIL", "切换唤醒词已复跑完整回环：小爱同学 -> 天猫精灵 -> 默认", [switch_next.step_dir, switch_to_tmall.step_dir, switch_idle_tmall_open.step_dir, switch_tmall_open.step_dir, switch_back_default.step_dir])
        evidence_map["SWAKE-007"] = [switch_next.step_dir, switch_to_tmall.step_dir, switch_idle_tmall_open.step_dir, switch_tmall_open.step_dir, switch_back_default.step_dir]

        runner.run_shell_step("assist_reg_config_clear", "config.clear", capture_s=3.0, ready_wait_s=1.0)
        runner.run_shell_step("assist_reg_reboot", "reboot", capture_s=10.0, ready_wait_s=8.0)

        reg_entry = runner.run_voice_sequence("reg_entry_timeout_wait30s", ["小度小度", "学习命令词"], post_wait_s=3.0)
        timeout_wait = runner.run_idle_wait_step("reg_entry_timeout_idle_30s", duration_s=30.0)
        reg_entry_recover = runner.run_voice_sequence("reg_entry_timeout_recover_open", ["小度小度", "打开电风扇"], post_wait_s=3.0)
        reg_entry_status = "PASS" if step_pass(reg_entry_recover, require_proto=True) else "FAIL"
        runner.add_case_result("REG-ENTRY-003", "语音注册", reg_entry_status, "学习入口超时后状态回收正常，默认控制链路恢复", [reg_entry.step_dir, timeout_wait.step_dir, reg_entry_recover.step_dir])
        evidence_map["REG-ENTRY-003"] = [reg_entry.step_dir, timeout_wait.step_dir, reg_entry_recover.step_dir]
        reg_cmd_learn = runner.run_voice_sequence("reg_cmd_learn_close_sequence", ["小度小度", "学习命令词", "学习下一个", "笑逐颜开", "笑逐颜开"], post_wait_s=12.0)
        reg_cmd_save_closure = runner.run_voice_sequence("reg_cmd_learn_close_save_closure", ["小度小度", "笑逐颜开"], post_wait_s=8.0)
        reg_cmd_reboot_after_save = runner.run_shell_step("reg_cmd_reboot_after_save", "reboot", capture_s=10.0, ready_wait_s=8.0)
        reg_cmd_alias = runner.run_voice_sequence("reg_cmd_alias_close_recheck", ["小度小度", "笑逐颜开"], post_wait_s=4.0)
        reg_cmd_default = runner.run_voice_sequence("reg_cmd_default_close_recheck", ["小度小度", "关闭电风扇"], post_wait_s=4.0)
        reg_cmd_save_text = reg_cmd_learn.log_text + "\n" + reg_cmd_save_closure.log_text
        reg_cmd_save_ok = text_has_any(reg_cmd_save_text, ["save new voice.bin", "reg cmd over success", "save config success"])
        reg_cmd_alias_ok = evidence_has_frame(reg_cmd_alias, words["关闭电风扇"]["发送协议"])
        reg_cmd_default_ok = evidence_has_frame(reg_cmd_default, words["关闭电风扇"]["发送协议"])
        reg_cmd_003_status = "PASS" if reg_cmd_save_ok and reg_cmd_alias_ok and reg_cmd_default_ok else "BLOCKED"
        runner.add_case_result(
            "REG-CMD-003",
            "语音注册-命令词",
            reg_cmd_003_status,
            (
                "学习保存闭环后重启，学习别名与原始默认命令均可触发关闭电风扇"
                if reg_cmd_003_status == "PASS"
                else f"命令词共存证据不足：save_ok={reg_cmd_save_ok}, alias_ok={reg_cmd_alias_ok}, default_ok={reg_cmd_default_ok}"
            ),
            [reg_cmd_learn.step_dir, reg_cmd_save_closure.step_dir, reg_cmd_reboot_after_save.step_dir, reg_cmd_alias.step_dir, reg_cmd_default.step_dir],
            {"save_ok": reg_cmd_save_ok, "alias_frames": proto_frames_from_hex(reg_cmd_alias.proto_hex), "default_frames": proto_frames_from_hex(reg_cmd_default.proto_hex)},
        )
        evidence_map["REG-CMD-003"] = [reg_cmd_learn.step_dir, reg_cmd_save_closure.step_dir, reg_cmd_reboot_after_save.step_dir, reg_cmd_alias.step_dir, reg_cmd_default.step_dir]

        runner.run_shell_step("assist_reg_wake_config_clear", "config.clear", capture_s=3.0, ready_wait_s=1.0)
        runner.run_shell_step("assist_reg_wake_reboot", "reboot", capture_s=10.0, ready_wait_s=8.0)

        reg_wake_learn = runner.run_voice_sequence("reg_wake_learn_sequence", ["小度小度", "学习唤醒词", "晴空万里", "晴空万里"], post_wait_s=3.0)
        reg_wake_template_full = runner.run_voice_sequence("reg_wake_template_full_reenter", ["小度小度", "学习唤醒词"], post_wait_s=3.0)
        reg_wake_open = runner.run_voice_sequence("reg_wake_newword_open", ["晴空万里", "打开电风扇"], post_wait_s=3.0)
        reg_wake_volume = runner.run_voice_sequence("reg_wake_newword_volume", ["晴空万里", "大声点"], post_wait_s=3.0)
        reg_wake_idle = runner.run_idle_wait_step("reg_wake_idle_before_other_ordinary_blocked", duration_s=22.0)
        reg_wake_other = runner.run_voice_sequence("reg_wake_other_ordinary_blocked", ["小爱同学", "打开电风扇"], post_wait_s=3.0)
        runner.add_case_result("REG-WAKE-003", "语音注册-唤醒词", "PASS" if step_pass(reg_wake_open, require_proto=True) and step_pass(reg_wake_volume, require_proto=True) else "FAIL", "学习唤醒词后，基础控制与音量交互都正常", [reg_wake_learn.step_dir, reg_wake_open.step_dir, reg_wake_volume.step_dir])
        evidence_map["REG-WAKE-003"] = [reg_wake_learn.step_dir, reg_wake_open.step_dir, reg_wake_volume.step_dir]

        runner.add_case_result("REG-WAKE-004", "语音注册-唤醒词", "PASS" if step_pass(reg_wake_other, require_proto=False) else "FAIL", "学习唤醒词后，其他普通唤醒词未误唤醒", [reg_wake_learn.step_dir, reg_wake_idle.step_dir, reg_wake_other.step_dir])
        evidence_map["REG-WAKE-004"] = [reg_wake_learn.step_dir, reg_wake_idle.step_dir, reg_wake_other.step_dir]

        runner.run_shell_step("assist_powerloss_config_clear", "config.clear", capture_s=3.0, ready_wait_s=1.0)
        runner.run_shell_step("assist_powerloss_reboot", "reboot", capture_s=10.0, ready_wait_s=8.0)
        reg_powerloss_entry = runner.run_voice_sequence("powerloss_during_case_enter_learning", ["小度小度", "学习命令词"], post_wait_s=2.0)
        reg_powerloss_boot = runner.run_powercycle_step("powerloss_during_case_boot", capture_s=10.0, ready_wait_s=8.0)
        reg_powerloss_recover = runner.run_voice_sequence("powerloss_during_case_recover_open", ["小度小度", "打开电风扇"], post_wait_s=3.0)
        runner.add_case_result("REG-SAVE-002", "语音注册-保持", "PASS" if step_pass(reg_powerloss_recover, require_proto=True) else "FAIL", "学习流程中突发断电后，设备仍能正常启动并恢复基础功能", [reg_powerloss_entry.step_dir, reg_powerloss_boot.step_dir, reg_powerloss_recover.step_dir])
        evidence_map["REG-SAVE-002"] = [reg_powerloss_entry.step_dir, reg_powerloss_boot.step_dir, reg_powerloss_recover.step_dir]

        runner.run_shell_step("assist_reg_cmd_retry_clear", "config.clear", capture_s=3.0, ready_wait_s=1.0)
        runner.run_shell_step("assist_reg_cmd_retry_reboot", "reboot", capture_s=10.0, ready_wait_s=8.0)
        reg_cmd_retry = runner.run_voice_sequence("reg_cfg_cmd_retry_exhaust_sequence", ["小度小度", "学习命令词", "学习下一个", "万事大吉", "心想事成", "打开电风扇", "笑逐颜开"], post_wait_s=12.0)
        reg_cmd_retry_probe = runner.run_voice_sequence("reg_cfg_cmd_retry_exhaust_failed_alias_probe", ["小度小度", "万事大吉"], post_wait_s=4.0)
        reg_cmd_retry_text = reg_cmd_retry.log_text + "\n" + reg_cmd_retry_probe.log_text
        reg_cmd_retry_cap = f"error cnt > {voice_reg['command_retry_count']}" in reg_cmd_retry_text
        reg_cmd_retry_failed = "reg failed!" in reg_cmd_retry_text
        reg_cmd_retry_no_alias = step_pass(reg_cmd_retry_probe, require_proto=False)
        reg_cmd_retry_status = "PASS" if reg_cmd_retry_cap and reg_cmd_retry_failed and reg_cmd_retry_no_alias else "BLOCKED"
        runner.add_case_result(
            "REG-CFG-003",
            "配置一致性-语音注册",
            reg_cmd_retry_status,
            (
                f"命令词失败耗尽出现上限与失败收口，失败后别名未触发控制协议；`reg simila error!` 捕获次数={count_occurrences(reg_cmd_retry_text, 'reg simila error!')}"
                if reg_cmd_retry_status == "PASS"
                else f"命令词失败耗尽证据不足：cap={reg_cmd_retry_cap}, failed={reg_cmd_retry_failed}, no_alias={reg_cmd_retry_no_alias}"
            ),
            [reg_cmd_retry.step_dir, reg_cmd_retry_probe.step_dir],
            {"retry_count_observed": count_occurrences(reg_cmd_retry_text, "reg simila error!"), "cap_marker": reg_cmd_retry_cap, "failed_marker": reg_cmd_retry_failed, "probe_frames": proto_frames_from_hex(reg_cmd_retry_probe.proto_hex)},
        )
        evidence_map["REG-CFG-003"] = [reg_cmd_retry.step_dir, reg_cmd_retry_probe.step_dir]

        runner.run_shell_step("assist_reg_wake_retry_clear", "config.clear", capture_s=3.0, ready_wait_s=1.0)
        runner.run_shell_step("assist_reg_wake_retry_reboot", "reboot", capture_s=10.0, ready_wait_s=8.0)
        reg_wake_retry = runner.run_voice_sequence("reg_cfg_wake_retry_exhaust_sequence", ["小度小度", "学习唤醒词", "小熊维尼", "小树小树", "小度小度", "晴空万里"], post_wait_s=12.0)
        reg_wake_retry_probe = runner.run_voice_sequence("reg_cfg_wake_retry_exhaust_failed_wake_probe", ["小熊维尼", "打开电风扇"], post_wait_s=4.0)
        reg_wake_retry_default = runner.run_voice_sequence("reg_cfg_wake_retry_exhaust_default_wake_ok", ["小度小度", "打开电风扇"], post_wait_s=4.0)
        reg_wake_retry_text = reg_wake_retry.log_text + "\n" + reg_wake_retry_probe.log_text
        reg_wake_retry_cap = f"error cnt > {voice_reg['wakeup_retry_count']}" in reg_wake_retry_text
        reg_wake_retry_failed = "reg failed!" in reg_wake_retry_text
        reg_wake_retry_blocked = step_pass(reg_wake_retry_probe, require_proto=False)
        reg_wake_retry_default_ok = step_pass(reg_wake_retry_default, require_proto=True)
        reg_wake_retry_status = "PASS" if reg_wake_retry_cap and reg_wake_retry_failed and reg_wake_retry_blocked and reg_wake_retry_default_ok else "BLOCKED"
        runner.add_case_result(
            "REG-CFG-004",
            "配置一致性-语音注册",
            reg_wake_retry_status,
            (
                f"唤醒词失败耗尽出现上限与失败收口，失败词不生效且默认唤醒正常；`reg simila error!` 捕获次数={count_occurrences(reg_wake_retry_text, 'reg simila error!')}"
                if reg_wake_retry_status == "PASS"
                else f"唤醒词失败耗尽证据不足：cap={reg_wake_retry_cap}, failed={reg_wake_retry_failed}, blocked={reg_wake_retry_blocked}, default_ok={reg_wake_retry_default_ok}"
            ),
            [reg_wake_retry.step_dir, reg_wake_retry_probe.step_dir, reg_wake_retry_default.step_dir],
            {"retry_count_observed": count_occurrences(reg_wake_retry_text, "reg simila error!"), "cap_marker": reg_wake_retry_cap, "failed_marker": reg_wake_retry_failed, "probe_frames": proto_frames_from_hex(reg_wake_retry_probe.proto_hex), "default_frames": proto_frames_from_hex(reg_wake_retry_default.proto_hex)},
        )
        evidence_map["REG-CFG-004"] = [reg_wake_retry.step_dir, reg_wake_retry_probe.step_dir, reg_wake_retry_default.step_dir]

        runner.run_shell_step("assist_reg_conflict_clear", "config.clear", capture_s=3.0, ready_wait_s=1.0)
        runner.run_shell_step("assist_reg_conflict_reboot", "reboot", capture_s=10.0, ready_wait_s=8.0)
        reg_conflict_spoken_word = "大声点"
        reg_conflict_seq = runner.run_voice_sequence("reg_conflict_cmd_volume_word_sequence", ["小度小度", "学习命令词", reg_conflict_spoken_word, reg_conflict_spoken_word], post_wait_s=10.0)
        reg_conflict_reboot = runner.run_shell_step("reg_conflict_reboot_after_attempt", "reboot", capture_s=10.0, ready_wait_s=8.0)
        reg_conflict_recheck = runner.run_voice_sequence("reg_conflict_cmd_volume_word_recheck", ["小度小度", reg_conflict_spoken_word], post_wait_s=4.0)
        volume_up_proto = words["大声点"].get("发送协议", "")
        saved_conflict = text_has_any(reg_conflict_seq.log_text, ["save new voice.bin", "save config success", "reg cmd over success"])
        volume_recheck_ok = bool(volume_up_proto) and evidence_has_frames(reg_conflict_recheck, [words["小度小度"]["发送协议"], volume_up_proto])
        if saved_conflict:
            reg_conflict_status = "FAIL"
        elif volume_recheck_ok:
            reg_conflict_status = "PASS"
        else:
            reg_conflict_status = "BLOCKED"
        runner.add_case_result(
            "REG-CONFLICT-001",
            "语音注册-冲突词",
            reg_conflict_status,
            (
                f"功能词 `{reg_conflict_spoken_word}` 未被保存为命令词，重启后原音量增大功能正常"
                if reg_conflict_status == "PASS"
                else (
                    f"功能词 `{reg_conflict_spoken_word}` 出现保存收口，疑似被错误学习为命令词"
                    if reg_conflict_status == "FAIL"
                    else f"功能词 `{reg_conflict_spoken_word}` 未见保存收口，但重启后音量功能回测证据不足"
                )
            ),
            [reg_conflict_seq.step_dir, reg_conflict_reboot.step_dir, reg_conflict_recheck.step_dir],
            {"tested_spoken_word": reg_conflict_spoken_word, "saved_conflict": saved_conflict, "recheck_frames": proto_frames_from_hex(reg_conflict_recheck.proto_hex), "expected_volume_up_proto": volume_up_proto},
        )
        evidence_map["REG-CONFLICT-001"] = [reg_conflict_seq.step_dir, reg_conflict_reboot.step_dir, reg_conflict_recheck.step_dir]

        runner.run_shell_step("final_config_clear", "config.clear", capture_s=3.0, ready_wait_s=1.0)
        runner.run_shell_step("final_reboot_clean", "reboot", capture_s=10.0, ready_wait_s=8.0)

        cfg_proto_1_expected = [words["小度小度"]["发送协议"], words["打开电风扇"]["发送协议"]]
        cfg_proto_1_status = "PASS" if evidence_has_frames(gate_interaction, cfg_proto_1_expected) else "FAIL"
        runner.add_case_result(
            "CFG-PROTO-001",
            "配置一致性-协议",
            cfg_proto_1_status,
            f"默认唤醒 + 打开电风扇协议链路={proto_frames_from_hex(gate_interaction.proto_hex)}",
            [gate_interaction.step_dir],
            {"frames": proto_frames_from_hex(gate_interaction.proto_hex), "expected_frames": cfg_proto_1_expected, "source": "testability_gate_default_wake_open"},
        )
        evidence_map["CFG-PROTO-001"] = [gate_interaction.step_dir]

        cfg_proto_2_status = "PASS" if evidence_has_frames(voice_off, [words["小度小度"]["发送协议"], words["关闭语音"]["发送协议"]]) and valid_proto.detail.get("payload_hex") == words["开语音"]["接收协议"] else "FAIL"
        runner.add_case_result(
            "CFG-PROTO-002",
            "配置一致性-协议",
            cfg_proto_2_status,
            f"关语音主动协议链路={proto_frames_from_hex(voice_off.proto_hex)}，开语音发送值={valid_proto.detail.get('payload_hex')}",
            [voice_off.step_dir, valid_proto.step_dir],
            {"voice_off_frames": proto_frames_from_hex(voice_off.proto_hex), "voice_on_payload": valid_proto.detail.get("payload_hex")},
        )
        evidence_map["CFG-PROTO-002"] = [voice_off.step_dir, valid_proto.step_dir]

        reg_cmd_repeat_actual = count_occurrences(reg_cmd_learn.log_text, "reg again!")
        reg_cmd_repeat_status = "PASS" if reg_cmd_repeat_actual == voice_reg["command_repeat_count"] - 1 else "FAIL"
        runner.add_case_result(
            "REG-CFG-001",
            "配置一致性-语音注册",
            reg_cmd_repeat_status,
            f"命令词学习过程中 `reg again!` 次数={reg_cmd_repeat_actual}，需求={voice_reg['command_repeat_count'] - 1}",
            [reg_cmd_learn.step_dir],
        )
        evidence_map["REG-CFG-001"] = [reg_cmd_learn.step_dir]

        reg_wake_repeat_actual = count_occurrences(reg_wake_learn.log_text, "reg again!")
        reg_wake_repeat_status = "PASS" if reg_wake_repeat_actual == voice_reg["wakeup_repeat_count"] - 1 else "FAIL"
        runner.add_case_result(
            "REG-CFG-002",
            "配置一致性-语音注册",
            reg_wake_repeat_status,
            f"唤醒词学习过程中 `reg again!` 次数={reg_wake_repeat_actual}，需求={voice_reg['wakeup_repeat_count'] - 1}",
            [reg_wake_learn.step_dir],
        )
        evidence_map["REG-CFG-002"] = [reg_wake_learn.step_dir]

        reg_cmd_template_full_marker = has_all_markers(reg_cmd_full_entry.log_text, ["reg over!", "play id : 34"])
        reg_cmd_template_status = "PASS" if reg_cmd_template_full_marker else "BLOCKED"
        runner.add_case_result(
            "REG-CFG-005",
            "配置一致性-语音注册",
            reg_cmd_template_status,
            (
                "命令词模板已满提示成立"
                if reg_cmd_template_status == "PASS"
                else "命令词模板数上限不能再用启动 `regCmdCount` 直接推断；需在本用例内主动填满两个命令模板并看到保存闭环后再判定，当前前置未闭合"
            ),
            [startup.step_dir, reg_cmd_full_entry.step_dir],
            {"boot_config": startup_cfg, "template_full_marker": reg_cmd_template_full_marker, "validation_note": "startup regCmdCount is auxiliary only, not a template-full precondition"},
        )
        evidence_map["REG-CFG-005"] = [startup.step_dir, reg_cmd_full_entry.step_dir]

        reg_wake_template_status = "PASS" if has_all_markers(reg_wake_template_full.log_text, ["reg over!", "play id : 34"]) and step_pass(reg_wake_open, require_proto=True) else "FAIL"
        runner.add_case_result(
            "REG-CFG-006",
            "配置一致性-语音注册",
            reg_wake_template_status,
            "学成一个唤醒词后再次进入学习唤醒词，已触发模板已满提示且原学习唤醒词仍有效",
            [reg_wake_learn.step_dir, reg_wake_template_full.step_dir, reg_wake_open.step_dir],
        )
        evidence_map["REG-CFG-006"] = [reg_wake_learn.step_dir, reg_wake_template_full.step_dir, reg_wake_open.step_dir]
    except UntestableFirmware as exc:
        runner.log_event("testability_gate_abort", {"reason": str(exc)})
    finally:
        runner.save_streams()
        runner.close_ports()

    update_case_markdown(runner.case_results, evidence_map)
    export_cases()
    shutil.copy2(PLAN_PATH, runner.static_dir / "plan" / PLAN_PATH.name)
    shutil.copy2(CASE_MD_PATH, runner.static_dir / "cases" / CASE_MD_PATH.name)
    shutil.copy2(CASE_XLSX_PATH, runner.static_dir / "cases" / CASE_XLSX_PATH.name)
    runner.write_case_results()
    runner.write_failure_analysis()
    runner.sync_bundle_root_artifacts()
    print(runner.bundle_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
