#!/usr/bin/env python
import argparse
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADER = [
    "章节",
    "用例ID",
    "功能模块",
    "用例类型",
    "测试点",
    "前置条件",
    "测试步骤",
    "主断言（功能）",
    "辅断言（协议 / 播报 / 日志）",
    "当前状态",
    "证据 / 备注",
]

STATUS_FILL = {
    "PASS": "C6EFCE",
    "FAIL": "FFC7CE",
    "TODO": "FFF2CC",
    "BLOCKED": "D9EAF7",
}


def split_row(line: str) -> list[str]:
    raw = line.strip()
    if not raw.startswith("|") or not raw.endswith("|"):
        return []
    cells = [cell.strip() for cell in raw[1:-1].split("|")]
    return [cell.replace("<br>", "\n") for cell in cells]


def is_separator_row(cells: list[str]) -> bool:
    if not cells:
        return False
    return all(set(cell) <= {"-", ":", " "} for cell in cells)


def parse_markdown_tables(path: Path) -> list[dict]:
    lines = path.read_text(encoding="utf-8").splitlines()
    current_section = ""
    rows: list[dict] = []
    in_table = False
    header_seen = False

    for line in lines:
        if line.startswith("### "):
            current_section = line[4:].strip()
            in_table = False
            header_seen = False
            continue

        if not line.startswith("|"):
            in_table = False
            header_seen = False
            continue

        cells = split_row(line)
        if not cells or is_separator_row(cells):
            continue

        if not in_table:
            in_table = True
            header_seen = True
            continue

        if header_seen and cells and cells[0] == "用例ID":
            continue

        if len(cells) != 10:
            continue

        rows.append(
            {
                "章节": current_section,
                "用例ID": cells[0].strip("`"),
                "功能模块": cells[1],
                "用例类型": cells[2],
                "测试点": cells[3],
                "前置条件": cells[4],
                "测试步骤": cells[5],
                "主断言（功能）": cells[6],
                "辅断言（协议 / 播报 / 日志）": cells[7],
                "当前状态": cells[8].strip("`"),
                "证据 / 备注": cells[9],
            }
        )

    return rows


def autosize(ws) -> None:
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            longest = max((len(part) for part in value.splitlines()), default=0)
            widths[cell.column_letter] = min(max(widths.get(cell.column_letter, 0), longest + 2), 60)
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build_workbook(rows: list[dict], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "正式用例"
    ws.freeze_panes = "A2"

    ws.append(HEADER)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for item in rows:
        ws.append([item[key] for key in HEADER])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        status = row[9].value
        fill = STATUS_FILL.get(status)
        if fill:
            row[9].fill = PatternFill("solid", fgColor=fill)

    autosize(ws)

    summary = wb.create_sheet("统计")
    summary.append(["统计项", "值"])
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)

    status_counter = Counter(item["当前状态"] for item in rows)
    module_counter = Counter(item["功能模块"] for item in rows)

    summary_rows = [
        ("总用例数", len(rows)),
        ("PASS", status_counter.get("PASS", 0)),
        ("FAIL", status_counter.get("FAIL", 0)),
        ("TODO", status_counter.get("TODO", 0)),
        ("BLOCKED", status_counter.get("BLOCKED", 0)),
        ("", ""),
        ("模块分布", ""),
    ]
    summary_rows.extend(module_counter.items())
    for row in summary_rows:
        summary.append(list(row))

    for row in summary.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    autosize(summary)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Export markdown case tables to an Excel workbook.")
    parser.add_argument("--input", required=True, help="Markdown case file path")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    rows = parse_markdown_tables(input_path)
    if not rows:
        raise SystemExit(f"No case rows parsed from: {input_path}")
    build_workbook(rows, output_path)
    print(f"exported {len(rows)} rows -> {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
