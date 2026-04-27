#!/usr/bin/env python
from __future__ import annotations

import os
import re
import subprocess
import sys
from collections import Counter
import os
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
REQ_DIR = Path(os.environ.get("TRISOLARIS_REQ_DIR", ROOT / "项目需求" / "CSK5062小度风扇需求")).expanduser().resolve()
DELIVERABLE_ROOT = ROOT / "deliverables" / "csk5062_xiaodu_fan"
PLAN_PATH = DELIVERABLE_ROOT / "plan" / "测试方案.md"
CASE_MD_PATH = DELIVERABLE_ROOT / "archive" / "测试用例-正式版.md"
CASE_XLSX_PATH = DELIVERABLE_ROOT / "cases" / "测试用例-正式版.xlsx"
EXPORT_SCRIPT = Path(__file__).resolve().parent / "export_case_md_to_xlsx.py"

TRUTH = {
    "default_wake": "小度小度",
    "switch_mode": "单次循环切换",
    "switch_persist": "是",
    "report_persist": "否",
    "voice_persist": "是",
    "voice_on_proto": "A5 FB 0A CC",
    "passive_proto": "A5 FB 12 CC",
}

EXEC_PROTO_PORT = os.environ.get("TRISOLARIS_PROTO_PORT", "").strip()
EXEC_LOG_PORT = os.environ.get("TRISOLARIS_LOG_PORT", "").strip()
EXEC_CTRL_PORT = os.environ.get("TRISOLARIS_CTRL_PORT", "").strip()
EXEC_DEVICE_KEY = os.environ.get("TRISOLARIS_DEVICE_KEY", "").strip()


def is_yes(value: str) -> bool:
    return value.strip() in {"是", "支持", "保存", "需要", "true", "True", "YES", "Yes", "yes", "1"}


def parse_requirement_markdown(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8")

    def expect_int(pattern: str) -> int:
        match = re.search(pattern, text)
        if not match:
            raise RuntimeError(f"Unable to parse int with pattern: {pattern}")
        return int(match.group(1))

    def expect_text(pattern: str) -> str:
        match = re.search(pattern, text)
        if not match:
            raise RuntimeError(f"Unable to parse text with pattern: {pattern}")
        return match.group(1).strip()

    return {
        "project_name": expect_text(r"项目名称：([^\n]+)"),
        "branch_name": expect_text(r"分支名称：([^\n]+)"),
        "chip": expect_text(r"芯片型号：([^\n]+)"),
        "version": expect_text(r"固件版本：([^\n]+)"),
        "wake_timeout_s": expect_int(r"唤醒时长:\s*(\d+)s"),
        "volume_steps": expect_int(r"音量档位:\s*(\d+)"),
        "default_volume": expect_int(r"初始化默认音量:\s*(\d+)"),
        "volume_underflow_tone": expect_text(r"最小音量下溢提示播报:\s*([^\n]+)"),
        "volume_overflow_tone": expect_text(r"最大音量上溢提示播报:\s*([^\n]+)"),
        "mic_analog_gain_db": expect_int(r"mic模拟增益:\s*(\d+)"),
        "mic_digital_gain_db": expect_int(r"mic数字增益:\s*(\d+)"),
        "proto_baud": expect_int(r"协议串口:\s*UART1、波特率(\d+)"),
        "log_baud": expect_int(r"日志串口:\s*UART0、波特率(\d+)"),
        "wake_power_save_raw": expect_text(r"唤醒词掉电保存:\s*([^\n]+)"),
        "volume_power_save_raw": expect_text(r"音量掉电保存:\s*([^\n]+)"),
    }


def load_word_table(path: Path) -> tuple[list[dict[str, str]], dict[str, dict[str, str]]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[1]]
    headers = ["" if cell.value is None else str(cell.value) for cell in sheet[1]]
    items: list[dict[str, str]] = []
    by_key: dict[str, dict[str, str]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        item = {headers[idx]: "" if row[idx] is None else str(row[idx]) for idx in range(len(headers))}
        items.append(item)
        semantic = item.get("语义(最小功能词)", "")
        func_type = item.get("功能类型", "")
        if semantic:
            by_key[semantic] = item
        if func_type and func_type not in by_key:
            by_key[func_type] = item
    return items, by_key


def load_voice_reg_config(path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows: list[tuple[str, str, Any]] = []
    category = ""
    for row in sheet.iter_rows(values_only=True):
        if not row or len(row) < 4 or row[2] is None:
            continue
        if row[1]:
            category = str(row[1]).strip()
        rows.append((category, str(row[2]).strip(), row[3]))

    def find_value(cat: str, key: str) -> Any:
        for row_cat, row_key, row_value in rows:
            if row_cat == cat and row_key == key:
                return row_value
        raise RuntimeError(f"Unable to find voice-reg config: {cat} / {key}")

    return {
        "command_mode": str(find_value("自学习种类及模式", "命令词学习模式")),
        "wakeup_repeat_count": int(find_value("自学习唤醒词参数", "自学习时每个词需说几遍")),
        "wakeup_word_max": int(find_value("自学习唤醒词参数", "自学习唤醒词字数上限")),
        "wakeup_word_min": int(find_value("自学习唤醒词参数", "自学习唤醒词字数下限")),
        "wakeup_template_count": int(find_value("自学习唤醒词参数", "自学习唤醒词模板数")),
        "wakeup_retry_count": int(find_value("自学习唤醒词参数", "唤醒词学习失败重试次数")),
        "command_repeat_count": int(find_value("自学习命令词参数", "自学习时每个词需说几遍")),
        "command_word_max": int(find_value("自学习命令词参数", "自学习命令词字数上限")),
        "command_word_min": int(find_value("自学习命令词参数", "自学习命令词字数下限")),
        "command_template_count": int(find_value("自学习命令词参数", "自学习命令词模板数")),
        "command_retry_count": int(find_value("自学习命令词参数", "命令词学习失败重试次数")),
    }


def seed_from_case_markdown(text: str) -> dict[str, dict[str, str]]:
    seed: dict[str, dict[str, str]] = {}
    for line in text.splitlines():
        if not line.startswith("| `"):
            continue
        cells = [cell.strip() for cell in line[1:-1].split("|")]
        if len(cells) != 10:
            continue
        seed[cells[0].strip("`")] = {"status": cells[8].strip("`"), "evidence": cells[9]}
    return seed


def capture_section(text: str, title: str) -> str:
    match = re.search(rf"(?ms)^### {re.escape(title)}\n.*?(?=^### |\n## |\Z)", text)
    if not match:
        raise RuntimeError(f"Missing section: {title}")
    return match.group(0).rstrip()


def table_block(title: str, rows: list[list[str]], seed: dict[str, dict[str, str]]) -> str:
    lines = [
        f"### {title}",
        "",
        "| 用例ID | 功能模块 | 用例类型 | 测试点 | 前置条件 | 测试步骤 | 主断言（功能） | 辅断言（协议 / 播报 / 日志） | 当前状态 | 证据 / 备注 |",
        "| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for row in rows:
        case_id = row[0]
        state = seed.get(case_id, {}).get("status", "TODO")
        evidence = seed.get(case_id, {}).get("evidence", "待执行后回填")
        lines.append(
            "| " + " | ".join([f"`{case_id}`", *row[1:], f"`{state}`", evidence]) + " |"
        )
    return "\n".join(lines)


def get_proto(words: dict[str, dict[str, str]], key: str, field: str, fallback: str = "") -> str:
    return words.get(key, {}).get(field, fallback)


def execution_mapping_lines(req: dict[str, Any]) -> list[str]:
    lines: list[str] = []
    if EXEC_PROTO_PORT:
        lines.append(f"- 本轮执行协议口映射：`{EXEC_PROTO_PORT} @ {req['proto_baud']}`")
    if EXEC_LOG_PORT:
        lines.append(f"- 本轮执行日志口映射：`{EXEC_LOG_PORT} @ {req['log_baud']}`")
    if EXEC_CTRL_PORT:
        lines.append(f"- 本轮执行控制 / boot 口映射：`{EXEC_CTRL_PORT}`")
    if EXEC_DEVICE_KEY:
        lines.append(f"- 本轮执行播报声卡：`{EXEC_DEVICE_KEY}`")
    if lines:
        lines.append("- 说明：以上仅是本轮工位设备映射，不改变需求里的 UART0 / UART1 逻辑职责。")
    return lines


def build_dynamic_blocks(req: dict[str, Any], reg: dict[str, Any], words: dict[str, dict[str, str]], wake_words: list[str]) -> dict[str, list[list[str]]]:
    default_wake = TRUTH["default_wake"] if TRUTH["default_wake"] in wake_words else wake_words[0]
    alt_wakes = [word for word in wake_words if word != default_wake]
    current_wake = alt_wakes[0] if alt_wakes else default_wake
    other_wake = alt_wakes[1] if len(alt_wakes) > 1 else default_wake
    volume_persist = is_yes(req["volume_power_save_raw"])
    wake_send = get_proto(words, default_wake, "发送协议")
    open_fan_send = get_proto(words, "打开电风扇", "发送协议")
    voice_off_send = get_proto(words, "关闭语音", "发送协议", "A5 FA 11 BB")
    voice_on_recv = get_proto(words, "开语音", "接收协议", TRUTH["voice_on_proto"])
    volume_power_title = (
        "音量设置应掉电保存（需求保持断电前档位）"
        if volume_persist
        else f"音量设置不应掉电保存（需求默认音量 `{req['default_volume']}`）"
    )
    volume_power_assert = (
        "重启后应保持断电前音量档位，不应恢复成其他值"
        if volume_persist
        else "重启后应恢复需求默认音量，而不是保留断电前档位"
    )
    volume_power_aux = (
        "`COM38` 启动日志中的 `volume` 与重启后行为应保持断电前档位"
        if volume_persist
        else "`COM38` 启动日志中的 `volume` 与重启后行为应回到默认值"
    )

    return {
        "4.1 环境与会话控制": [
            ["ENV-001", "环境确认", "正例", "当前固件身份确认", "设备已启动，日志口可读", "1. 读取启动日志<br>2. 核对项目名、版本号、关键配置", "当前运行固件与目标风扇项目一致", f"`COM38` 版本字段与目标固件 `{req['version']}` 一致"],
            ["ENV-002", "烧录确认", "正例", "烧录后版本闭环", "完成一次烧录或重烧", "1. 重启设备<br>2. 抓版本日志并比对", "可确认当前设备确实运行待测固件", f"`COM38` 启动日志显示 `version: {req['version'].replace('V', '')}`"],
            ["SESS-001", "启动待机", "正例", "上电欢迎语与待机就绪", "设备断电后重新上电", "1. 上电启动<br>2. 观察欢迎语和待机收口", "设备正常进入可交互待机态", "欢迎语、启动日志完整"],
            ["SESS-002", "默认唤醒", "正例", "默认唤醒词可进入会话", "设备处于待机态", f"1. 说 `{default_wake}`<br>2. 跟随说 `打开电风扇`", "默认唤醒词可进入会话并完成后续交互", f"`COM36` 出现唤醒协议 `{wake_send}` 和开风扇协议 `{open_fan_send}`"],
            ["SESS-003", "会话超时", "正例", f"唤醒超时退出与休息语（需求 `{req['wake_timeout_s']}s`）", "设备处于待机态", f"1. 仅说 `{default_wake}`<br>2. 静默等待超时", "会话自动退出，不再维持交互态", "`COM38` 出现 `TIME_OUT`、`MODE=0`、休息语 `play id`"],
            ["SESS-004", "会话超时", "反例", "超时退出后未重唤醒时命令阻断", "已发生一次超时退出", "1. 超时退出后不重唤醒<br>2. 直接说 `打开电风扇`", "后续命令不应直接生效", "`COM36` 不应出现新的主动控制协议"],
            ["SESS-005", "退出识别", "正例", "退出识别后会话结束", "已进入交互态", "1. 说 `退出识别`<br>2. 不重唤醒直接说控制词", "退出后后续命令被阻断", "`COM36` 出现退出识别协议，随后无新控制协议"],
            ["SESS-006", "退出识别", "交叉", "退出识别后重新唤醒恢复正常", "已执行一次退出识别", f"1. 退出识别<br>2. 再次说 `{default_wake} -> 打开电风扇`", "重新唤醒后功能恢复正常", "可辅助看再次出现唤醒协议和控制协议"],
            ["SESS-007", "待机阻断", "反例", "未唤醒状态直说命令不应生效", "设备处于待机态", "1. 不先唤醒<br>2. 直接说 `打开电风扇`", "不应直接执行控制功能", "`COM36` 不应出现主动控制协议"],
        ],
        "4.2 基础控制与音量": [
            ["CTRL-001", "基础控制", "正例", "打开电风扇", "已通过任一有效唤醒词进入会话", "1. 说 `打开电风扇`", "风扇打开功能成立", f"`COM36` 出现对应控制协议 `{open_fan_send}`，`COM38` 命中正确词"],
            ["CTRL-002", "基础控制", "正例", "关闭电风扇", "已进入会话", "1. 说 `关闭电风扇`", "风扇关闭功能成立", "`COM36` 出现对应控制协议"],
            ["CTRL-003", "基础控制", "正例", "开机", "已进入会话", "1. 说 `开机`", "开机功能成立", "`COM36` 出现对应控制协议"],
            ["CTRL-004", "基础控制", "正例", "关机", "已进入会话", "1. 说 `关机`", "关机功能成立", "`COM36` 出现对应控制协议"],
            ["VOL-001", "音量控制", "正例", "大小声调节", "已进入会话", "1. 说 `大声点`<br>2. 再说 `小声点`", "音量可按方向增减", "`COM36` 协议与词命中正确"],
            ["VOL-002", "音量控制", "正例", "最大 / 最小边界", "已进入会话", "1. 说 `最大音量`<br>2. 说 `最小音量`", "边界设置成立，边界反馈正确", f"边界提示应分别为 `{req['volume_overflow_tone']}` / `{req['volume_underflow_tone']}`"],
            ["VOL-003", "音量控制", "配置一致性", volume_power_title, "设备可正常调音并支持断电重启", "1. 将音量改到非默认档位<br>2. 断电重启<br>3. 观察启动配置并复测音量状态", volume_power_assert, volume_power_aux],
            ["VOL-004", "音量控制", "异常", "连续快速调音稳定性", "已进入会话", "1. 连续快速执行音量词<br>2. 观察档位与提示", "不应出现跳档、反跳或状态错乱", "协议序列、日志档位与提示应一致"],
        ],
        "4.4 切换唤醒词及交叉验证": [
            ["SWAKE-001", "切换唤醒词", "正例", f"当前切换唤醒词 `{current_wake}` 立即生效", f"默认唤醒词 `{default_wake}` 可用", f"1. 说 `切换唤醒词`<br>2. 用当前目标词 `{current_wake}` 唤醒", "当前切换唤醒词可唤醒", "切换提示 tone 与日志匹配"],
            ["SWAKE-002", "切换唤醒词", "正例", f"默认唤醒词 `{default_wake}` 仍常驻", "已完成一次切换", f"1. 使用默认词 `{default_wake}` 再次唤醒", "默认唤醒词不因切换而失效", "可辅助看再次出现唤醒链路"],
            ["SWAKE-003", "切换唤醒词", "反例", f"非当前非默认唤醒词 `{other_wake}` 不能误唤醒", "已完成一次切换", f"1. 用其他候选唤醒词 `{other_wake}` 尝试唤醒", "非当前非默认词不应误唤醒", "`COM36` 不应出现新的主动唤醒 / 控制协议"],
            ["SWAKE-004", "切换唤醒词", "交叉", "切换后用当前唤醒词执行基础控制与音量调节", f"已切换到目标唤醒词 `{current_wake}`", f"1. 用 `{current_wake}` 唤醒<br>2. 说 `打开电风扇`<br>3. 再说 `大声点`", "当前唤醒词下，基础控制和音量交互都正常", "辅助看唤醒协议、控制协议、音量协议"],
            ["SWAKE-005", "切换唤醒词", "交叉", "切换后用默认唤醒词执行基础控制与音量调节", f"已切换到目标唤醒词 `{current_wake}`", f"1. 用 `{default_wake}` 唤醒<br>2. 说 `打开电风扇`<br>3. 再说 `大声点`", "默认唤醒词下，正常交互仍成立", "可辅助看控制与音量协议"],
            ["SWAKE-006", "切换唤醒词", "正例", f"当前切换结果掉电保持（需求 `{TRUTH['switch_persist']}`）", "已切换成功并看到保存收口", f"1. 掉电重启<br>2. 用 `{current_wake}` 和 `{default_wake}` 复测", "当前切换结果跨掉电保持，默认词仍常驻", "启动配置与复测行为一致"],
            ["SWAKE-007", "切换唤醒词", "异常", f"多轮切换回环稳定性（{TRUTH['switch_mode']}）", f"候选唤醒词列表已知：`{' / '.join(wake_words)}`", "1. 连续多次执行 `切换唤醒词`<br>2. 走完整轮回环", "切换顺序稳定，无越界、无多词并行生效", "可辅助看每轮提示与实际生效矩阵"],
        ],
        "4.7 配置与参数一致性专项": [
            ["CFG-AUDIO-001", "配置一致性-音频", "正例", "mic 增益 / 输入配置与需求一致", "设备已启动，日志口可读", "1. 上电启动<br>2. 在启动日志中查找 `AADC: AGAIN` 与 `DGAIN`", f"模拟增益应为 `{req['mic_analog_gain_db']}dB`，数字增益应为 `{req['mic_digital_gain_db']}dB`", f"辅助看 `AGAIN={req['mic_analog_gain_db']}dB`、`DGAIN={req['mic_digital_gain_db']}dB`"],
            ["CFG-WAKE-001", "配置一致性-会话参数", "正例", f"唤醒会话时长应为 `{req['wake_timeout_s']}s`", "设备处于待机态", f"1. 说 `{default_wake}`<br>2. 静默等待超时<br>3. 记录唤醒到超时回收的实际间隔", f"会话超时应接近并满足需求 `{req['wake_timeout_s']}s`", "`COM38` 辅助看 `MODE=1`、`TIME_OUT`、`MODE=0` 及回收时间"],
            ["CFG-VOL-001", "配置一致性-音量参数", "正例", f"默认音量应为 `{req['default_volume']}`", "烧录前已执行 `config.clear -> reboot -> burn`，设备处于烧录后默认状态", "1. 读取烧录后首启 `Running Config` 的 `volume` 字段<br>2. 从当前默认音量开始单边连续执行 `大声点` 到最大边界并统计有效 step<br>3. 再通过最大到最小、最小到最大双向探测拿到全部档位<br>4. 用 `总档位数 - 默认到最大有效 step` 推导默认档位", f"探测推导出的默认音量档位应为 `{req['default_volume']}`，启动配置字段应作为零基/内部档位辅助交叉验证", "`COM38` 看 `Running Config`、`mini player set vol`、边界 `play id`、`refresh config volume=`；默认值不只按配置字段单点判定"],
            ["CFG-VOL-002", "配置一致性-音量参数", "正例", f"音量总档位应为 `{req['volume_steps']}`", "设备可正常调音", "1. 从最小档开始逐级调大音量<br>2. 记录所有可达档位及边界行为", f"应存在连续 `{req['volume_steps']}` 个有效档位", "可辅助看 `refresh config volume=`、边界 tone 与控制效果"],
            ["CFG-PROTO-001", "配置一致性-协议", "正例", "默认唤醒与基础控制主动协议一致性", "协议串口 `COM36` 可抓包", f"1. 执行 `{default_wake} -> 打开电风扇`<br>2. 读取 `COM36` 原始协议", f"应出现唤醒协议 `{wake_send}` 与开风扇协议 `{open_fan_send}`", "`COM38` 仅辅助看识别链路是否完整"],
            ["CFG-PROTO-002", "配置一致性-协议", "正例", "开关语音协议一致性", "已确认语音入口功能存在", f"1. 说 `关闭语音`<br>2. `COM36` 发送 `{voice_on_recv}`<br>3. 复测基础交互", f"关语音主动协议应为 `{voice_off_send}`，开语音接收协议应为 `{voice_on_recv}`", "`COM38` 辅助看 `receive msg` 与状态刷新"],
            ["CFG-PROTO-003", "配置一致性-协议", "正例", "被动播报协议一致性", "协议串口可发送 hex", f"1. `COM36` 发送 `{TRUTH['passive_proto']}`<br>2. 观察被动播报链路", f"协议 `{TRUTH['passive_proto']}` 应能触发被动播报", "`COM38` 辅助看 `receive msg:: ...`、`play start`、`play id`、`play stop`"],
            ["REG-CFG-001", "配置一致性-语音注册", "正例", f"命令词学习次数 = `{reg['command_repeat_count']}`", "已进入命令词学习态", "1. 执行一轮命令词学习<br>2. 统计学习过程中的再次录入提示次数", f"应出现 `{reg['command_repeat_count'] - 1}` 次再次录入提示，对应总学习次数 `{reg['command_repeat_count']}`", "辅助看 `reg again!` 次数与学习收口"],
            ["REG-CFG-002", "配置一致性-语音注册", "正例", f"唤醒词学习次数 = `{reg['wakeup_repeat_count']}`", "已进入唤醒词学习态", "1. 执行一轮唤醒词学习<br>2. 统计学习过程中的再次录入提示次数", f"应出现 `{reg['wakeup_repeat_count'] - 1}` 次再次录入提示，对应总学习次数 `{reg['wakeup_repeat_count']}`", "辅助看 `reg again!` 次数与学习收口"],
            ["REG-CFG-003", "配置一致性-语音注册", "正例", f"命令词失败重试上限 = `{reg['command_retry_count']}`", "已进入命令词学习态", "1. 连续输入错误样本直到失败<br>2. 统计失败重试次数", f"错误重试次数应为 `{reg['command_retry_count']}`，随后明确失败退出", "辅助看 `reg simila error!`、`error cnt > N`、`reg failed!`"],
            ["REG-CFG-004", "配置一致性-语音注册", "正例", f"唤醒词失败重试上限 = `{reg['wakeup_retry_count']}`", "已进入唤醒词学习态", "1. 连续输入错误样本直到失败<br>2. 统计失败重试次数", f"错误重试次数应为 `{reg['wakeup_retry_count']}`，随后明确失败退出", "辅助看 `reg simila error!`、`error cnt > N`、`reg failed!`"],
            ["REG-CFG-005", "配置一致性-语音注册", "正例", f"命令词模板数上限 = `{reg['command_template_count']}`", "本用例内已主动学满命令词模板，并观察到保存闭环", "1. 连续学满命令词模板<br>2. 保存闭环后重启<br>3. 再次进入对应学习流程", f"达到 `{reg['command_template_count']}` 个模板后，应出现模板已满 / 学习结束提示", "不能用启动 `regCmdCount` 单点推断模板已满；必须有主动填满、保存、重入三段证据"],
            ["REG-CFG-006", "配置一致性-语音注册", "正例", f"唤醒词模板数上限 = `{reg['wakeup_template_count']}`", "已完成唤醒词学习并继续尝试学习", "1. 学满唤醒词模板<br>2. 再次进入唤醒词学习流程", f"达到 `{reg['wakeup_template_count']}` 个模板后，应出现模板已满 / 学习结束提示", "辅助看 `reg over!` 与重入学习时的阻断表现"],
        ],
    }


def build_plan_markdown(req: dict[str, Any], reg: dict[str, Any], wake_words: list[str]) -> str:
    default_wake = TRUTH["default_wake"] if TRUTH["default_wake"] in wake_words else wake_words[0]
    volume_persist = is_yes(req["volume_power_save_raw"])
    volume_power_desc = "掉电保持" if volume_persist else "掉电不保持"
    runtime_mapping = execution_mapping_lines(req)
    runtime_block = ("\n" + "\n".join(runtime_mapping)) if runtime_mapping else ""
    return f"""# {req['project_name']}测试方案

## 1. 方案目标

本方案根据当前需求文档与用户最新口径动态生成，用于验证 `{req['project_name']}` 的功能、参数和异常行为。

统一按 3 层判定：

1. 功能是否正常、使能状态是否正确
2. 数值 / 参数 / 协议是否与需求一致
3. 异常场景下是否稳定，不出现重启、卡死、无法继续交互或部分功能失效

## 2. 输入依据与项目真值

- 项目：`{req['project_name']}`
- 分支：`{req['branch_name']}`
- 芯片：`{req['chip']}`
- 固件版本：`{req['version']}`
- 默认唤醒词：`{default_wake}`
- 候选唤醒词：`{' / '.join(wake_words)}`
- 唤醒时长：`{req['wake_timeout_s']}s`
- 音量档位：`{req['volume_steps']}`
- 默认音量：`{req['default_volume']}`
- mic 增益：`{req['mic_analog_gain_db']} / {req['mic_digital_gain_db']}`
- 协议串口：`COM36 @ {req['proto_baud']}`
- 日志串口：`COM38 @ {req['log_baud']}`
- 播报声卡：`{EXEC_DEVICE_KEY or 'VID_8765&PID_5678:8_804B35B_1_0000'}`{runtime_block}
- 唤醒词掉电保存原始字段：`{req['wake_power_save_raw']}`
- 音量掉电保存原始字段：`{req['volume_power_save_raw']}`
- 用户确认：切换唤醒词单次循环切换且结果掉电保存；关播报不掉电保存；开关语音掉电保存；开语音需走协议 `{TRUTH['voice_on_proto']}`

## 3. 执行与判定规则

- 先过“烧录后健康检查 + 最小可测性门禁”，再进入正式需求验证
- 若设备支持全配置清除指令，烧录前先执行 `config.clear -> reboot -> burn`；默认音量判定必须结合烧录后首启配置和单边/双边音量探测，避免旧配置或零基字段造成误判
- 先验证功能主链路，再验证数值 / 协议，再验证异常场景
- 一条用例只回答一个问题；状态建立写在前置条件，主体步骤只负责判定该问题是否成立
- 对当前项目重点看 5 类状态依赖：是否已唤醒、当前唤醒词、语音开关状态、当前音量档位、学习/掉电前保存状态
- 协议结论以 `COM36` 为准，`COM38` 只做辅助日志
- 证据分层：协议口原始抓取 / 注入与功能行为属于必要证据，日志 / 播报 / 配置刷新属于辅助证据
- 保存类结论必须先看到保存完成日志
- 功能成立但参数不一致时，继续后续回归，并把差异单独记为配置一致性问题
- 每次播报要等上一条响应结束；默认间隔 `4.5s`，若已检测到 `play stop` 可提前继续
- 断电保持要验证“断电前功能生效 + 保存完成 + 重启后仍保持”
- 断电不保持要验证“断电前功能生效 + 重启后恢复默认值”

## 4. 分模块验证思路

### 4.1 环境与会话

- 验证烧录闭环、版本闭环、上电待机、默认唤醒、超时退出、退出识别和待机阻断
- 会话功能既要验证能否正常收口，也要验证超时值是否真为 `{req['wake_timeout_s']}s`

### 4.2 基础控制与音量

- 验证打开/关闭电风扇、开机/关机
- 验证音量增减、边界提示、默认音量、总档位数和{volume_power_desc}；默认音量用“默认点到边界 step + 全档位探测”推导实际档位

### 4.3 播报与语音开关

- 关播报：仅抑制播报，不影响识别、控制和协议
- 关语音：阻断所有语音入口
- 开语音：通过协议 `{TRUTH['voice_on_proto']}` 恢复
- 开关语音还要验证错误协议反例与掉电保持

### 4.4 切换唤醒词与语音注册

- 切换唤醒词后要验证当前词、默认词、其他词阻断和交叉交互
- 语音注册要验证入口、学习成功、重启保持、删除、冲突、失败重试、模板上限和数值一致性

## 5. 重点收敛项

1. 唤醒会话时长是否真为 `{req['wake_timeout_s']}s`
2. 默认音量是否真为 `{req['default_volume']}`（必须先清配置烧录，再用探测法推导默认档位）
3. 音量总档位是否真为 `{req['volume_steps']}`
4. 音量掉电行为是否与需求一致（当前需求：`{req['volume_power_save_raw']}`）
5. 开关语音协议和掉电保持是否与项目真值一致
6. 语音注册学习次数、失败重试上限、模板数是否与配置一致
"""


def render_case_markdown(existing_text: str, req: dict[str, Any], reg: dict[str, Any], words: dict[str, dict[str, str]], wake_words: list[str]) -> str:
    seed = seed_from_case_markdown(existing_text)
    dynamic = build_dynamic_blocks(req, reg, words, wake_words)
    keep_43 = capture_section(existing_text, "4.3 播报与语音开关")
    keep_45 = capture_section(existing_text, "4.5 语音注册入口、成功与保持")
    keep_46 = capture_section(existing_text, "4.6 语音注册失败、冲突与删除")
    runtime_mapping = execution_mapping_lines(req)
    device_key = EXEC_DEVICE_KEY or "VID_8765&PID_5678:8_804B35B_1_0000"

    parts = [
        f"# {req['project_name']}正式测试用例",
        "",
        "## 1. 编写说明",
        "",
        "- 本文档依据 `需求文档.md`、`词条处理.xlsx`、`tone.h`、`语音注册功能.xlsx` 以及当前用户确认口径动态生成。",
        "- 本文档把“测试方案”继续拆成可执行的正式用例，保留功能主线、交叉验证、反例、异常和参数一致性专项。",
        "- 当前项目统一采用“功能主判定、参数独立判定、协议 / 播报 / 日志辅助定位”的结论口径。",
        "- 用例编排遵循“门禁先行 -> 功能验证 -> 参数一致性 -> 异常 / 边界”的顺序，不把数值问题混成纯功能 FAIL。",
        "- 当前项目的关键状态依赖包括：唤醒态、当前唤醒词、语音开关、音量档位、学习态以及掉电前是否已完成保存。",
        "- 当前状态字段说明：",
        "  - `PASS`：已有有效实测证据闭环",
        "  - `FAIL`：已有有效实测证据证明当前固件不满足需求",
        "  - `TODO`：需求和方案已明确，但当前保留为人工验证或待后续补证据",
        "  - `BLOCKED`：当前前置状态或环境不足，暂无法形成结论",
        "",
        "## 2. 用例字段说明",
        "",
        "- `用例ID`：唯一标识",
        "- `功能模块`：该用例属于哪个功能模块",
        "- `用例类型`：正例 / 反例 / 异常 / 交叉 / 配置一致性",
        "- `测试点`：这条用例到底在验证什么",
        "- `前置条件`：执行前必须满足的状态",
        "- `测试步骤`：怎么测",
        "- `主断言（功能）`：功能是否真正发生 / 不发生，这是主结论来源",
        "- `辅断言（协议 / 播报 / 日志）`：用于辅助定位，不代替主结论",
        "- `当前状态`：`PASS` / `FAIL` / `TODO` / `BLOCKED`",
        "- `证据 / 备注`：对齐到现有 `result/` 或 `deliverables/` 证据目录，未执行则写待回填",
        "",
        "## 3. 固定环境与执行约束",
        "",
        f"- 固件：`{req['version']}`",
        f"- 协议串口：`COM36 @ {req['proto_baud']}`",
        f"- 日志串口：`COM38 @ {req['log_baud']}`",
        "- 电源 / boot 控制口：`COM39`",
        f"- 播报声卡：`{device_key}`",
        "- 串口使用约束：`COM36` / `COM38` 不能并发占用，协议窗口和自动化抓取必须串行使用",
        "- 语料执行约束：",
        "  - 上一条响应播报未收口前，不播下一条语料",
        "  - 默认按 `4.5s` 节奏推进；若已检测到 `play stop` / `quiet_after_response`，可提前继续",
        "- 保存类统一约束：看到保存完成日志后，才允许对掉电保持下结论",
        "- 烧录前基线约束：若设备支持 `config.clear` 或等效全配置清除指令，先执行 `config.clear -> reboot -> burn`，确保烧录后首启使用固件默认设置",
        "",
        "## 4. 正式测试用例",
        "",
        table_block("4.1 环境与会话控制", dynamic["4.1 环境与会话控制"], seed),
        "",
        table_block("4.2 基础控制与音量", dynamic["4.2 基础控制与音量"], seed),
        "",
        keep_43,
        "",
        table_block("4.4 切换唤醒词及交叉验证", dynamic["4.4 切换唤醒词及交叉验证"], seed),
        "",
        keep_45,
        "",
        keep_46,
        "",
        table_block("4.7 配置与参数一致性专项", dynamic["4.7 配置与参数一致性专项"], seed),
        "",
    ]
    if runtime_mapping:
        insert_at = parts.index("## 4. 正式测试用例")
        parts[insert_at:insert_at] = [*runtime_mapping, ""]

    rendered = "\n".join(parts) + "\n"
    statuses: dict[str, str] = {}
    for line in rendered.splitlines():
        if not line.startswith("| `"):
            continue
        cells = [cell.strip() for cell in line[1:-1].split("|")]
        if len(cells) != 10:
            continue
        statuses[cells[0].strip("`")] = cells[8].strip("`")
    counter = Counter(statuses.values())
    fail_ids = [case_id for case_id, status in statuses.items() if status == "FAIL"]
    todo_ids = [case_id for case_id, status in statuses.items() if status == "TODO"]
    section_counts = {
        "4.1 环境与会话控制": 9,
        "4.2 基础控制与音量": 8,
        "4.3 播报与语音开关": 9,
        "4.4 切换唤醒词及交叉验证": 7,
        "4.5 语音注册入口、成功与保持": 14,
        "4.6 语音注册失败、冲突与删除": 12,
        "4.7 配置与参数一致性专项": 13,
    }
    rendered += "## 5. 当前统计\n\n"
    rendered += f"- 当前正式用例总数：`{sum(section_counts.values())}`\n"
    rendered += "- 模块分布：\n"
    for title, count in section_counts.items():
        rendered += f"  - `{title}`：`{count}`\n"
    rendered += f"- 当前状态分布：`PASS={counter.get('PASS', 0)}`、`FAIL={counter.get('FAIL', 0)}`、`BLOCKED={counter.get('BLOCKED', 0)}`、`TODO={counter.get('TODO', 0)}`\n"
    rendered += "- 当前明确保留缺陷：" + ("无" if not fail_ids else "、".join(f"`{case_id}`" for case_id in fail_ids)) + "\n\n"
    rendered += "## 6. 当前使用建议\n\n"
    rendered += "- 后续若需求文档发生变化，应先重跑“需求解析 -> 方案生成 -> 用例生成”链路，再执行自动化或半自动化验证。\n"
    rendered += "- 评审“需求覆盖度”时，优先看第 4 章是否已经覆盖功能主线、交叉验证、反例、异常和参数专项。\n"
    rendered += "- 评审“执行结果”时，优先结合最新时间戳报告目录中的 `case_results.json`、`failure_analysis.md` 和串口日志一起看。\n"
    rendered += "- 当前仅剩人工保留项 `TODO` 为：\n"
    if todo_ids:
        for case_id in todo_ids:
            rendered += f"  - `{case_id}`\n"
    else:
        rendered += "  - 无\n"
    return rendered


def main() -> int:
    req = parse_requirement_markdown(REQ_DIR / "需求文档.md")
    word_items, words = load_word_table(REQ_DIR / "词条处理.xlsx")
    reg = load_voice_reg_config(REQ_DIR / "语音注册功能.xlsx")
    wake_words = [item["语义(最小功能词)"] for item in word_items if item.get("功能类型") == "唤醒词" and item.get("语义(最小功能词)")]
    existing_case_text = CASE_MD_PATH.read_text(encoding="utf-8", errors="replace") if CASE_MD_PATH.exists() else ""

    PLAN_PATH.parent.mkdir(parents=True, exist_ok=True)
    CASE_MD_PATH.parent.mkdir(parents=True, exist_ok=True)
    CASE_XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)

    PLAN_PATH.write_text(build_plan_markdown(req, reg, wake_words), encoding="utf-8")
    CASE_MD_PATH.write_text(render_case_markdown(existing_case_text, req, reg, words, wake_words), encoding="utf-8")
    subprocess.run([sys.executable, str(EXPORT_SCRIPT), "--input", str(CASE_MD_PATH), "--output", str(CASE_XLSX_PATH)], cwd=ROOT, check=True)
    print(f"generated: {PLAN_PATH}")
    print(f"generated: {CASE_MD_PATH}")
    print(f"generated: {CASE_XLSX_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
